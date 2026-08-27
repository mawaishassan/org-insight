import os
import uuid
import logging
import asyncio
from datetime import datetime
from typing import Any
from openpyxl import load_workbook
from sqlalchemy import select, insert, delete, func, text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.core.config import get_settings
from app.core.models import KpiBulkUploadTask, KPIEntry, KPIField, KpiMultiLineRow, KpiMultiLineCell, FieldType
from app.entries.service import (
    get_reference_allowed_values,
    _normalize_reference_value,
    _is_reference_empty_or_sentinel,
    coerce_multi_reference_raw,
    filter_multi_reference_to_allowed,
    _is_subfield_satisfied_for_row,
    coerce_mixed_list_raw,
)
from app.entries.routes import (
    _stringify_for_upsert_match_key,
    _is_multi_items_row_effectively_empty,
    _multi_line_cell_insert_row,
    _resolve_multi_items_import_mode,
    _upsert_merge_multi_line_items,
    mark_entry_modified,
    propagate_formula_recalculations,
)

logger = logging.getLogger(__name__)

def utc_now():
    return datetime.utcnow()

def _parse_excel_file_sync(temp_file_path: str, sub_fields_info: list[dict]) -> list[dict]:
    """Load and parse the Excel file synchronously in a separate worker thread."""
    wb = load_workbook(filename=temp_file_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            raise Exception("The Excel sheet is empty")
        
        def _norm_header(x: object) -> str:
            return str(x).strip() if x is not None else ""
        
        header = [_norm_header(x) for x in header_row]
        key_to_idx = {k: i for i, k in enumerate(header) if k}
        
        allowed_subfields_keys = {s["key"] for s in sub_fields_info}
        allowed_subfields_by_key = {s["key"]: s for s in sub_fields_info}
        allowed_lower = {str(s["key"]).strip().lower(): s["key"] for s in sub_fields_info}
        name_to_key = {s["name"].strip(): s["key"] for s in sub_fields_info if s.get("name")}
        name_to_key_lower = {str(s["name"]).strip().lower(): s["key"] for s in sub_fields_info if s.get("name")}
        
        def resolve_col_to_key(col: str) -> str | None:
            col_s = str(col or "").strip()
            if not col_s:
                return None
            if col_s in allowed_subfields_keys:
                return col_s
            if col_s in name_to_key:
                return name_to_key[col_s]
            c_low = col_s.lower()
            if c_low in allowed_lower:
                return allowed_lower[c_low]
            if c_low in name_to_key_lower:
                return name_to_key_lower[c_low]
            return None

        raw_rows = []
        for r in row_iter:
            if r is None:
                continue
            item = {}
            empty = True
            for col, idx in key_to_idx.items():
                key = resolve_col_to_key(col)
                if not key:
                    continue
                raw = r[idx] if idx < len(r) else None
                if raw is None or raw == "":
                    continue
                empty = False
                sf = allowed_subfields_by_key[key]
                sf_type = sf["field_type"]
                if sf_type in ("attachment", "formula"):
                    continue
                elif sf_type == "number":
                    try:
                        item[key] = float(raw)
                    except Exception:
                        item[key] = str(raw)
                elif sf_type == "boolean":
                    if isinstance(raw, bool):
                        item[key] = raw
                    else:
                        s = str(raw).strip().lower()
                        item[key] = s in ("1", "true", "yes", "y")
                elif sf_type == "date":
                    if hasattr(raw, "date"):
                        try:
                            item[key] = raw.date().isoformat()
                        except Exception:
                            item[key] = str(raw)
                    else:
                        item[key] = str(raw)
                elif sf_type == "multi_reference":
                    item[key] = str(raw).strip() if raw is not None else ""
                elif sf_type == "mixed_list":
                    item[key] = coerce_mixed_list_raw(str(raw) if raw is not None else "") or None
                else:
                    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
                        item[key] = _stringify_for_upsert_match_key(raw)
                    else:
                        item[key] = str(raw) if raw is not None else ""
            
            if not empty and not _is_multi_items_row_effectively_empty(item):
                raw_rows.append(item)
        return raw_rows
    finally:
        wb.close()


async def process_bulk_upload_task(
    task_id: str,
    temp_file_path: str,
    entry_id: int,
    field_id: int,
    org_id: int | None,
    import_mode: str,
    match_sub_field_key: str | None,
    current_user_id: int,
    session_factory: async_sessionmaker[AsyncSession],
):
    """Asynchronous background worker task for Multi-Line Items Excel bulk import."""
    logger.info("Starting bulk upload task: %s for entry_id=%s, field_id=%s", task_id, entry_id, field_id)
    
    settings = get_settings()
    db: AsyncSession = session_factory()
    
    try:
        # Step 1: Update task status to PARSING
        await _update_task(db, task_id, status="PARSING", progress_percent=0.0)
        
        # Load field definition to know the sub-fields structure
        field_res = await db.execute(
            select(KPIField).where(KPIField.id == field_id)
        )
        field = field_res.scalar_one_or_none()
        if not field:
            raise Exception("Multi-line item field not found")
        
        # Prepare sub_fields info list to avoid accessing lazy properties in another thread
        sub_fields_info = [
            {
                "key": getattr(sf, "key", ""),
                "name": getattr(sf, "name", ""),
                "field_type": sf.field_type.value if hasattr(sf.field_type, "value") else str(sf.field_type),
            }
            for sf in (field.sub_fields or [])
        ]
        
        # Parse Excel using threadpool to avoid blocking main asyncio loop
        raw_rows = await asyncio.to_thread(
            _parse_excel_file_sync,
            temp_file_path,
            sub_fields_info,
        )
            
        total_rows = len(raw_rows)
        logger.info("Found %s rows to import", total_rows)
        await _update_task(db, task_id, status="VALIDATING", total_rows=total_rows, progress_percent=5.0)
        
        # Load entry to check details
        entry_res = await db.execute(
            select(KPIEntry).where(KPIEntry.id == entry_id)
        )
        entry = entry_res.scalar_one_or_none()
        if not entry:
            raise Exception("Entry not found")
            
        # Clean subfields conditions dependencies
        key_to_sf = {sf.key: sf for sf in (field.sub_fields or [])}
        for item in raw_rows:
            changed = True
            while changed:
                changed = False
                for sf in (field.sub_fields or []):
                    if item.get(sf.key) is not None:
                        if not _is_subfield_satisfied_for_row(sf, item, key_to_sf):
                            item[sf.key] = None
                            changed = True

        # Validation phase: check reference allowed values in cache
        validation_errors = []
        ref_sub_fields = [s for s in (field.sub_fields or []) if getattr(s, "field_type", None) == FieldType.reference]
        multif_sub_fields = [s for s in (field.sub_fields or []) if getattr(s, "field_type", None) == FieldType.multi_reference]
        
        seen_ck = set()
        ordered_ck = []
        for sf in (*ref_sub_fields, *multif_sub_fields):
            cfg = getattr(sf, "config", None) or {}
            sid = cfg.get("reference_source_kpi_id")
            skey = cfg.get("reference_source_field_key")
            subkey = cfg.get("reference_source_sub_field_key")
            if not sid or not skey:
                continue
            ck = (int(sid), str(skey), str(subkey) if subkey else None)
            if ck not in seen_ck:
                seen_ck.add(ck)
                ordered_ck.append(ck)
                
        ref_list_by_ck = {}
        if ordered_ck:
            async def _fetch_allowed(ck):
                sid, skey, subkey = ck
                lst = await get_reference_allowed_values(
                    db,
                    int(sid),
                    str(skey),
                    org_id,
                    source_sub_field_key=subkey,
                    year=entry.year,
                )
                return ck, lst
            ref_list_by_ck = dict(await asyncio.gather(*[_fetch_allowed(ck) for ck in ordered_ck]))
            
        allowed_cache = {
            ck: {_normalize_reference_value(a) for a in lst} for ck, lst in ref_list_by_ck.items()
        }
        
        for row_idx, item in enumerate(raw_rows):
            # Check references
            for sf in ref_sub_fields:
                cfg = getattr(sf, "config", None) or {}
                sid = cfg.get("reference_source_kpi_id")
                skey = cfg.get("reference_source_field_key")
                subkey = cfg.get("reference_source_sub_field_key")
                if not sid or not skey:
                    continue
                cache_key = (int(sid), str(skey), str(subkey) if subkey else None)
                cell = item.get(sf.key)
                raw = cell if isinstance(cell, str) else str(cell) if cell is not None else ""
                normalized = _normalize_reference_value(raw)
                if _is_reference_empty_or_sentinel(normalized):
                    item[sf.key] = None
                    continue
                if normalized not in allowed_cache.get(cache_key, set()):
                    validation_errors.append({
                        "row_index": row_idx + 2,  # spreadsheet format (1-based + 1 header)
                        "sub_field_name": sf.name,
                        "sub_field_key": sf.key,
                        "value": raw,
                        "message": "Value does not exist in the referenced KPI field."
                    })
                    
            # Check multi references
            for sf in multif_sub_fields:
                cfg = getattr(sf, "config", None) or {}
                sid = cfg.get("reference_source_kpi_id")
                skey = cfg.get("reference_source_field_key")
                subkey = cfg.get("reference_source_sub_field_key")
                if not sid or not skey:
                    continue
                cache_key = (int(sid), str(skey), str(subkey) if subkey else None)
                allowed_list = ref_list_by_ck.get(cache_key, [])
                allowed_norm = allowed_cache.get(cache_key, set())
                cell = item.get(sf.key)
                for tok in coerce_multi_reference_raw(cell):
                    if isinstance(tok, dict):
                        s = next((str(tok[k]) for k in ("label", "text", "value", "name") if k in tok and tok[k] is not None), "")
                    else:
                        s = str(tok) if tok is not None else ""
                    n = _normalize_reference_value(s)
                    if _is_reference_empty_or_sentinel(n):
                        continue
                    if n not in allowed_norm:
                        validation_errors.append({
                            "row_index": row_idx + 2,
                            "sub_field_name": sf.name,
                            "sub_field_key": sf.key,
                            "value": str(cell),
                            "message": "One or more values do not exist in the referenced KPI field."
                        })
                        break
                else:
                    cleaned = filter_multi_reference_to_allowed(cell, allowed_list) if allowed_list else []
                    item[sf.key] = cleaned if cleaned else None
                    
            # Update validation status incrementally for large lists
            if row_idx > 0 and row_idx % 25000 == 0:
                val_progress = 5.0 + (float(row_idx) / total_rows) * 15.0
                await _update_task(db, task_id, progress_percent=val_progress)
                
        if validation_errors:
            # Save validation errors and abort
            await _update_task(db, task_id, status="FAILED", progress_percent=100.0, validation_errors=validation_errors[:500]) # Cap errors returned to UI
            return
            
        # Step 3: Fast Database Insert/Upsert in chunks
        await _update_task(db, task_id, status="IMPORTING", progress_percent=20.0)
        
        mode = _resolve_multi_items_import_mode(import_mode, import_mode == "append")
        
        # Load existing items to merge if needed
        existing_rows_list = []
        if mode != "replace":
            from app.entries.routes import _load_multi_line_row_dicts
            existing_pairs = await _load_multi_line_row_dicts(db, entry_id=entry_id, field=field)
            existing_rows_list = [r for _, r in existing_pairs] if existing_pairs else []
            
        if mode == "replace":
            # Clear all current rows
            await db.execute(
                delete(KpiMultiLineRow).where(
                    KpiMultiLineRow.entry_id == entry_id,
                    KpiMultiLineRow.field_id == field_id,
                )
            )
            await db.flush()
            final_rows = raw_rows
            rows_added = total_rows
            rows_updated = 0
            rows_overridden = total_rows
        elif mode == "append":
            final_rows = existing_rows_list + raw_rows
            rows_added = total_rows
            rows_updated = 0
            rows_overridden = 0
        else: # upsert
            mk = (match_sub_field_key or "").strip()
            sub_by_key = {s.key: s for s in (field.sub_fields or [])}
            match_ft = getattr(sub_by_key[mk], "field_type", None) if mk in sub_by_key else None
            merged, rows_updated, rows_added = _upsert_merge_multi_line_items(
                existing_rows_list, raw_rows, mk, match_ft
            )
            final_rows = merged
            rows_overridden = 0
            
        # Write final rows list in chunks of 5000 rows
        write_chunk = 5000
        total_final_rows = len(final_rows)
        
        # Clear existing table if doing upsert/append since we rewrite from merged final_rows
        if mode != "replace" and final_rows:
            await db.execute(
                delete(KpiMultiLineRow).where(
                    KpiMultiLineRow.entry_id == entry_id,
                    KpiMultiLineRow.field_id == field_id,
                )
            )
            await db.flush()

        from app.entries.service import _resolve_attachment_filenames_to_urls
        resolved_rows = await _resolve_attachment_filenames_to_urls(db, field.kpi_id, final_rows, field.sub_fields or [])
        key_to_sub = {getattr(s, "key", None): s for s in (field.sub_fields or []) if getattr(s, "key", None)}
        
        # Turn Postgres local triggers search rebuild off
        is_sqlite = "sqlite" in settings.DATABASE_URL.lower()
        if not is_sqlite:
            await db.execute(text("SET LOCAL app.disable_search_rebuild = 'on';"))
            
        for chunk_start in range(0, total_final_rows, write_chunk):
            chunk_end = min(chunk_start + write_chunk, total_final_rows)
            chunk_rows = resolved_rows[chunk_start:chunk_end]
            
            ts_now = utc_now()
            row_data = [
                {
                    "entry_id": entry_id,
                    "field_id": field_id,
                    "row_index": chunk_start + idx,
                    "created_at": ts_now,
                    "updated_at": ts_now,
                }
                for idx in range(len(chunk_rows))
            ]
            
            # 1. Insert KpiMultiLineRow for this chunk
            await db.execute(insert(KpiMultiLineRow), row_data)
            await db.flush()
            
            # Get generated row IDs for this chunk
            res_ids = await db.execute(
                select(KpiMultiLineRow.id).where(
                    KpiMultiLineRow.entry_id == entry_id,
                    KpiMultiLineRow.field_id == field_id,
                    KpiMultiLineRow.row_index >= chunk_start,
                    KpiMultiLineRow.row_index < chunk_end
                ).order_by(KpiMultiLineRow.row_index)
            )
            ids_ordered = [int(x[0]) for x in res_ids.all()]
            
            # 2. Prepare cells rows for this chunk
            cell_rows = []
            for idx, rdict in enumerate(chunk_rows):
                mlr_id = ids_ordered[idx]
                for k, v in rdict.items():
                    sub = key_to_sub.get(k)
                    if not sub:
                        continue
                    if getattr(sub, "field_type", None) == FieldType.mixed_list:
                        v = coerce_mixed_list_raw(v) or None
                    m = _multi_line_cell_insert_row(mlr_id, sub, v)
                    m["created_at"] = ts_now
                    m["updated_at"] = ts_now
                    cell_rows.append(m)
                    
            # 3. Bulk Insert cell rows for this chunk
            cell_chunk = 8000
            for ci in range(0, len(cell_rows), cell_chunk):
                await db.execute(insert(KpiMultiLineCell), cell_rows[ci : ci + cell_chunk])
            await db.flush()
            
            # 4. Commit this chunk to free memory and unlock resources
            await db.commit()
            
            # 5. Update Progress percentage (from 20% to 90%)
            processed = chunk_end
            pct = 20.0 + (float(processed) / total_final_rows) * 70.0
            await _update_task(db, task_id, processed_rows=processed, progress_percent=pct)
            
        # Rebuild searchable text for Postgres (Step 4)
        if not is_sqlite and resolved_rows:
            await db.execute(
                text(
                    """
                    UPDATE kpi_multi_line_rows r
                    SET search_text = src.t
                    FROM (
                      SELECT
                        c.row_id,
                        lower(
                          string_agg(
                            coalesce(
                              c.value_text,
                              c.value_json::text,
                              c.value_number::text,
                              c.value_boolean::text,
                              c.value_date::text,
                              ''
                            ),
                            ' '
                          )
                        ) AS t
                      FROM kpi_multi_line_cells c
                      JOIN kpi_multi_line_rows r2 ON c.row_id = r2.id
                      WHERE r2.entry_id = :entry_id AND r2.field_id = :field_id
                      GROUP BY c.row_id
                    ) AS src
                    WHERE src.row_id = r.id AND r.entry_id = :entry_id AND r.field_id = :field_id
                    """
                ),
                {"entry_id": entry_id, "field_id": field_id},
            )
            await db.flush()
            await db.commit()

        # Recalculate formulas and complete task
        await _update_task(db, task_id, progress_percent=95.0)
        await mark_entry_modified(db, entry, current_user_id)
        await propagate_formula_recalculations(db, entry_id=entry_id, org_id=org_id)
        await db.commit()
        
        await _update_task(
            db, 
            task_id, 
            status="COMPLETED", 
            progress_percent=100.0,
            rows_added=rows_added,
            rows_updated=rows_updated,
            rows_overridden=rows_overridden
        )
        
    except Exception as e:
        logger.exception("Error processing bulk upload task %s", task_id)
        await db.rollback()
        await _update_task(db, task_id, status="FAILED", progress_percent=100.0, error_message=str(e))
    finally:
        await db.close()
        # Clean up temporary file
        try:
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
        except Exception:
            pass

async def _update_task(db: AsyncSession, task_id: str, **kwargs):
    """Utility helper to update bulk upload task progress inside DB."""
    # NOTE: KpiBulkUploadTask has no updated_at column — do NOT add one here.
    if kwargs.get("status") in ("COMPLETED", "FAILED"):
        kwargs["completed_at"] = utc_now()
    
    from sqlalchemy import update
    await db.execute(
        update(KpiBulkUploadTask)
        .where(KpiBulkUploadTask.id == task_id)
        .values(**kwargs)
    )
    await db.commit()
