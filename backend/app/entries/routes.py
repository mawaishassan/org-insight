"""KPI entry API routes (data entry + admin lock)."""

from __future__ import annotations

from io import BytesIO
from datetime import datetime
import csv
import json
from typing import Any
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, distinct, func, and_, cast, or_
from sqlalchemy.sql import nulls_last
from sqlalchemy.types import String
from sqlalchemy.orm import selectinload
from sqlalchemy.orm.attributes import flag_modified
import httpx

from app.core.database import get_db
from app.auth.dependencies import get_current_user, require_org_admin
from app.core.models import (
    User,
    KPIEntry,
    KPIField,
    KPIFieldSubField,
    KPIFieldValue,
    KPI,
    FieldType,
    Organization,
    TimeDimension,
    effective_kpi_time_dimension,
    KpiMultiLineRowAccess,
    KpiMultiLineRow,
    KpiMultiLineCell,
    KpiRoleAssignment,
    UserOrganizationRole,
)
from app.entries.schemas import EntryCreate, EntrySubmit, EntryLock, EntryResponse, FieldValueResponse
from app.entries.service import (
    get_or_create_entry,
    user_can_edit_kpi,
    user_can_view_kpi,
    get_user_field_access_for_kpi,
    user_can_view_field,
    user_can_edit_field,
    user_can_edit_multi_line_field,
    user_can_add_row_multi_line_field,
    user_can_edit_row,
    user_can_delete_row,
    save_entry_values,
    submit_entry,
    lock_entry,
    list_entries,
    list_available_kpis,
    list_entries_overview,
    EntryValidationError,
    _normalize_reference_value,
    _stringify_for_upsert_match_key,
    _upsert_merge_multi_line_items,
    _is_multi_items_row_effectively_empty,
    parse_upsert_match_keys_json,
    coerce_mixed_list_raw,
)
from app.fields.service import list_fields as list_kpi_fields_service
from app.kpis.service import sync_kpi_entry_from_api
from app.entries.multi_item_filters import row_passes_filters
from app.entries.multi_line_load import load_multi_line_row_dicts as _load_multi_line_row_dicts
from app.entries.reference_filter_resolve import build_reference_resolution_map

router = APIRouter(prefix="/entries", tags=["entries"])


def _org_id(user: User, org_id_param: int | None) -> int:
    if org_id_param is not None and user.role.value == "SUPER_ADMIN":
        return org_id_param
    if user.organization_id is not None:
        return user.organization_id
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Organization required")


async def _reindex_row_access_after_delete(
    db: AsyncSession,
    *,
    entry_id: int,
    field_id: int,
    deleted_indices: set[int],
) -> None:
    """
    Keep row-level access attached to the same logical rows after deleting rows from value_json.
    - Access rows for deleted indices are removed.
    - Access rows after deleted positions are shifted down by the number of removed rows before them.
    """
    if not deleted_indices:
        return
    rules_res = await db.execute(
        select(KpiMultiLineRowAccess).where(
            KpiMultiLineRowAccess.entry_id == entry_id,
            KpiMultiLineRowAccess.field_id == field_id,
        )
    )
    rules = list(rules_res.scalars().all())
    if not rules:
        return
    sorted_deleted = sorted(deleted_indices)
    to_shift: list[tuple[KpiMultiLineRowAccess, int]] = []
    for rule in rules:
        current = int(rule.row_index)
        if current in deleted_indices:
            await db.delete(rule)
            continue
        shift = 0
        for deleted_idx in sorted_deleted:
            if deleted_idx < current:
                shift += 1
        if shift > 0:
            to_shift.append((rule, current - shift))
    await db.flush()
    # Two-phase reindex avoids transient uniqueness collisions on
    # (user_id, entry_id, field_id, row_index) while rows are shifting.
    for idx, (rule, _) in enumerate(to_shift, start=1):
        rule.row_index = -idx
    await db.flush()
    for rule, final_index in to_shift:
        rule.row_index = final_index
    await db.flush()


async def _reindex_multi_line_rows(
    db: AsyncSession,
    *,
    entry_id: int,
    field_id: int,
) -> None:
    """
    Reindex relational multi-line rows so row_index is dense 0..N-1.
    Two-phase update avoids transient uniqueness collisions on (entry_id, field_id, row_index).
    """
    res = await db.execute(
        select(KpiMultiLineRow).where(
            KpiMultiLineRow.entry_id == entry_id,
            KpiMultiLineRow.field_id == field_id,
        ).order_by(KpiMultiLineRow.row_index)
    )
    rows = list(res.scalars().all())
    if not rows:
        return
    # Temp negative indices
    for i, r in enumerate(rows, start=1):
        r.row_index = -i
    await db.flush()
    # Final dense indices
    for i, r in enumerate(rows):
        r.row_index = i
    await db.flush()


async def _replace_multi_line_rows_from_dicts(
    db: AsyncSession,
    *,
    entry_id: int,
    field: KPIField,
    rows: list[dict],
) -> None:
    """Replace relational multi-line rows/cells for (entry, field) from legacy list-of-dicts."""
    # Delete existing rows (cells cascade)
    existing_res = await db.execute(
        select(KpiMultiLineRow).where(
            KpiMultiLineRow.entry_id == entry_id,
            KpiMultiLineRow.field_id == field.id,
        )
    )
    for r in list(existing_res.scalars().all()):
        await db.delete(r)
    await db.flush()

    key_to_sub = {getattr(s, "key", None): s for s in (field.sub_fields or []) if getattr(s, "key", None)}

    def _add_cell(row_id: int, sub: Any, raw_val: Any) -> None:
        c = KpiMultiLineCell(row_id=row_id, sub_field_id=int(getattr(sub, "id")))
        ft = getattr(sub, "field_type", None)
        ft_s = ft.value if hasattr(ft, "value") else str(ft)
        if raw_val is None:
            pass
        elif ft_s == "number":
            try:
                c.value_number = float(raw_val)
            except Exception:
                c.value_text = str(raw_val)
        elif ft_s == "boolean":
            if isinstance(raw_val, bool):
                c.value_boolean = raw_val
            else:
                s = str(raw_val).strip().lower()
                if s in ("true", "yes", "1"):
                    c.value_boolean = True
                elif s in ("false", "no", "0"):
                    c.value_boolean = False
                else:
                    c.value_text = str(raw_val)
        elif ft_s == "date":
            c.value_text = str(raw_val)
        elif ft_s in ("reference", "multi_reference", "mixed_list", "attachment"):
            if isinstance(raw_val, (dict, list)):
                c.value_json = raw_val
            else:
                c.value_text = str(raw_val)
        else:
            if isinstance(raw_val, (dict, list)):
                c.value_json = raw_val
            else:
                c.value_text = str(raw_val)
        db.add(c)

    for idx, row in enumerate(rows or []):
        rdict = row if isinstance(row, dict) else {}
        mlr = KpiMultiLineRow(entry_id=entry_id, field_id=field.id, row_index=int(idx))
        db.add(mlr)
        await db.flush()
        for k, v in rdict.items():
            sub = key_to_sub.get(k)
            if not sub:
                continue
            if getattr(sub, "field_type", None) == FieldType.mixed_list:
                v = coerce_mixed_list_raw(v) or None
            _add_cell(mlr.id, sub, v)


def _entry_to_response(entry):
    entered_by_name = None
    if getattr(entry, "user", None):
        u = entry.user
        entered_by_name = (getattr(u, "full_name", None) or getattr(u, "username", None) or "").strip() or getattr(u, "username", None)
    return EntryResponse(
        id=entry.id,
        kpi_id=entry.kpi_id,
        organization_id=entry.organization_id,
        user_id=entry.user_id,
        year=entry.year,
        period_key=getattr(entry, "period_key", "") or "",
        is_draft=entry.is_draft,
        is_locked=entry.is_locked,
        submitted_at=entry.submitted_at,
        values=[
            FieldValueResponse(
                field_id=fv.field_id,
                value_text=fv.value_text,
                value_number=fv.value_number,
                value_json=fv.value_json,
                value_boolean=fv.value_boolean,
                value_date=fv.value_date,
            )
            for fv in (entry.field_values or [])
        ],
        entered_by_user_name=entered_by_name,
        updated_at=getattr(entry, "updated_at", None),
    )


async def _load_multi_items_field(db: AsyncSession, org_id: int, field_id: int) -> KPIField | None:
    """Load a multi_line_items field with sub_fields, scoped to org."""
    res = await db.execute(
        select(KPIField)
        .join(KPI, KPI.id == KPIField.kpi_id)
        .where(KPIField.id == field_id, KPI.organization_id == org_id)
        .options(selectinload(KPIField.sub_fields))
    )
    field = res.scalar_one_or_none()
    if not field or field.field_type != FieldType.multi_line_items:
        return None
    return field


def _serialize_multi_item_cell_for_xlsx(val: Any, field_type: FieldType | str | None) -> Any:
    """Ensure cell values are openpyxl-safe (lists/dicts are flattened or JSON — never raw list)."""
    if val is None:
        return ""
    ft = field_type.value if isinstance(field_type, FieldType) else field_type
    if ft == FieldType.multi_reference.value or ft == "multi_reference":
        if isinstance(val, list):
            parts = [str(x).strip() for x in val if x is not None and str(x).strip() != ""]
            return "; ".join(parts)
        return str(val).strip() if str(val).strip() else ""
    if ft == FieldType.mixed_list.value or ft == "mixed_list":
        if isinstance(val, list):
            parts = [str(x).strip() for x in val if x is not None and str(x).strip() != ""]
            return "; ".join(parts)
        return str(val).strip() if val is not None and str(val).strip() else ""
    if isinstance(val, list):
        parts = [str(x).strip() for x in val if x is not None and str(x).strip() != ""]
        return "; ".join(parts)
    if isinstance(val, dict):
        return json.dumps(val, ensure_ascii=False)
    return val


def _serialize_multi_item_cell_for_csv(val: Any, field_type: FieldType | str | None) -> Any:
    if val is None:
        return ""
    ft = field_type.value if isinstance(field_type, FieldType) else field_type
    if ft == FieldType.multi_reference.value or ft == "multi_reference":
        if isinstance(val, list):
            return "; ".join(str(x).strip() for x in val if x is not None and str(x).strip() != "")
        return val
    if ft == FieldType.mixed_list.value or ft == "mixed_list":
        if isinstance(val, list):
            return "; ".join(str(x).strip() for x in val if x is not None and str(x).strip() != "")
        return val
    if isinstance(val, list):
        return "; ".join(str(x).strip() for x in val if x is not None and str(x).strip() != "")
    return val


def _resolve_multi_items_import_mode(import_mode: str | None, append_legacy: bool) -> str:
    if import_mode:
        return import_mode.strip().lower()
    return "append" if append_legacy else "replace"


def _xlsx_bytes_for_multi_items_template(field: KPIField, existing_items: list[dict] | None = None) -> bytes:
    """Create an Excel template for multi_line_items sub_fields. Optionally include existing data."""
    # Import here so app can start even if optional dependency isn't installed yet.
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    sub_fields = list(field.sub_fields or [])
    key_to_sf = {s.key: s for s in sub_fields}
    # Header row uses sub-field keys (stable API identifiers).
    keys = [s.key for s in sub_fields]
    ws.append(keys)
    items = existing_items if isinstance(existing_items, list) else []
    if items:
        for item in items:
            if not isinstance(item, dict):
                continue
            row_cells = []
            for k in keys:
                raw = item.get(k)
                v = raw if raw is not None else ""
                sf = key_to_sf.get(k)
                ft = getattr(sf, "field_type", None) if sf else None
                row_cells.append(_serialize_multi_item_cell_for_xlsx(v, ft))
            ws.append(row_cells)
    # No blank data row: avoids a phantom row on every re-upload; users add rows in Excel as needed.

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_multi_items_xlsx(content: bytes, field: KPIField) -> list[dict]:
    """Parse uploaded xlsx into list[dict[sub_key] = value] for the field's sub_fields."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    key_to_idx = {k: i for i, k in enumerate(header) if k}
    allowed = {s.key: s for s in (field.sub_fields or [])}

    # Accept either key or name as column header (keys preferred).
    name_to_key = {s.name.strip(): s.key for s in (field.sub_fields or []) if s.name}

    def resolve_col_to_key(col: str) -> str | None:
        if col in allowed:
            return col
        if col in name_to_key:
            return name_to_key[col]
        return None

    out: list[dict] = []
    for r in rows[1:]:
        if r is None:
            continue
        item: dict = {}
        empty = True
        for col, idx in key_to_idx.items():
            key = resolve_col_to_key(col)
            if not key:
                continue
            raw = r[idx] if idx < len(r) else None
            if raw is None or raw == "":
                continue
            empty = False
            sf = allowed[key]
            if sf.field_type == FieldType.number:
                try:
                    item[key] = float(raw)
                except Exception:
                    # keep as string if can't coerce
                    item[key] = str(raw)
            elif sf.field_type == FieldType.boolean:
                if isinstance(raw, bool):
                    item[key] = raw
                else:
                    s = str(raw).strip().lower()
                    item[key] = s in ("1", "true", "yes", "y")
            elif sf.field_type == FieldType.date:
                # Store ISO date string (matches UI expectation for <input type=\"date\">)
                if hasattr(raw, "date"):
                    try:
                        item[key] = raw.date().isoformat()
                    except Exception:
                        item[key] = str(raw)
                else:
                    item[key] = str(raw)
            elif sf.field_type == FieldType.multi_reference:
                # Semicolon (or comma) separated in Excel; validated on upload.
                item[key] = str(raw).strip() if raw is not None else ""
            elif sf.field_type == FieldType.mixed_list:
                # Semicolon separated values in Excel; infer number/date/string.
                item[key] = coerce_mixed_list_raw(str(raw) if raw is not None else "") or None
            else:
                # Text / reference / attachment: Excel often stores numeric ids as float (e.g. 42.0).
                if isinstance(raw, (int, float)) and not isinstance(raw, bool):
                    item[key] = _stringify_for_upsert_match_key(raw)
                else:
                    item[key] = str(raw) if raw is not None else ""
        if not empty and not _is_multi_items_row_effectively_empty(item):
            out.append(item)
    return out


class MultiItemsRow(BaseModel):
    index: int
    data: dict
    can_edit: bool = True
    can_delete: bool = True


class MultiItemsListResponse(BaseModel):
    total: int
    page: int
    page_size: int
    rows: list[MultiItemsRow]
    sub_fields: list[dict]


class MultiItemsPageContextSubField(BaseModel):
    id: int | None = None
    key: str
    name: str
    field_type: str | None = None
    is_required: bool | None = None
    sort_order: int | None = None
    config: dict | None = None
    can_view: bool | None = None
    can_edit: bool | None = None


class MultiItemsPageContextField(BaseModel):
    id: int
    kpi_id: int
    name: str
    key: str
    field_type: str
    full_page_multi_items: bool | None = None
    row_level_user_access_enabled: bool | None = None
    config: dict | None = None
    sub_fields: list[MultiItemsPageContextSubField] = []


class MultiItemsPageContextResponse(BaseModel):
    entry_id: int
    kpi_id: int
    kpi_name: str
    field: MultiItemsPageContextField
    can_edit: bool
    kpi_level_can_edit: bool
    can_add_row: bool


class EntryIdResponse(BaseModel):
    id: int
    created: bool = False


@router.get("/multi-items/template")
async def download_multi_items_template(
    field_id: int = Query(...),
    entry_id: int | None = Query(None, description="If provided, include existing rows for this entry/field"),
    include_existing_rows: bool = Query(
        False,
        description="When true and entry_id is provided, include existing rows in the Excel download (can be slow for large datasets).",
    ),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download an Excel template for a multi_line_items field (any user who can edit this KPI)."""
    org_id = _org_id(current_user, organization_id)
    field = await _load_multi_items_field(db, org_id, field_id)
    if not field:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")
    can_edit = await user_can_edit_kpi(db, current_user.id, field.kpi_id)
    if not can_edit:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to edit this KPI")

    existing_items: list[dict] | None = None
    if include_existing_rows and entry_id is not None:
        entry_res = await db.execute(
            select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
        )
        entry = entry_res.scalar_one_or_none()
        if entry and entry.kpi_id == field.kpi_id:
            pairs = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=field)
            existing_items = [r for _, r in pairs] if pairs else None

    content = _xlsx_bytes_for_multi_items_template(field, existing_items=existing_items)
    filename = f"multi_items_{field.key}_{field.id}.xlsx"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/multi-items/upload")
async def upload_multi_items_excel(
    entry_id: int = Query(...),
    field_id: int = Query(...),
    append: bool = Query(False, description="Deprecated: use import_mode=append instead"),
    import_mode: str | None = Query(
        None,
        description="replace (default), append, or upsert (requires match_sub_field_key)",
        pattern="^(replace|append|upsert)$",
    ),
    match_sub_field_key: str | None = Query(
        None,
        description="Sub-field key used to match rows when import_mode=upsert (same normalized value => update row)",
    ),
    file: UploadFile = File(...),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload Excel and set/append/upsert multi_line_items value_json for an entry+field (requires add_row permission)."""
    org_id = _org_id(current_user, organization_id)

    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entry not found")
    if entry.is_locked:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Entry is locked")

    field = await _load_multi_items_field(db, org_id, field_id)
    if field:
        can_add = await user_can_add_row_multi_line_field(db, current_user.id, field.kpi_id, field.id)
        if not can_add:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to add rows to this field")
    if not field:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")
    if field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Field does not belong to entry KPI")

    content = await file.read()
    items = _parse_multi_items_xlsx(content, field)
    items = [it for it in items if isinstance(it, dict) and not _is_multi_items_row_effectively_empty(it)]

    # Reference consistency check for reference sub-fields (row-level errors)
    # Rules:
    # - Missing/blank/"NA" values => coerce to null (do not validate)
    # - Invalid values => return row/column error and abort upload (no partial import)
    from app.entries.service import (
        get_reference_allowed_values,
        _normalize_reference_value,
        _is_reference_empty_or_sentinel,
        coerce_multi_reference_raw,
        filter_multi_reference_to_allowed,
    )
    validation_errors: list[dict] = []
    ref_sub_fields = [s for s in (field.sub_fields or []) if getattr(s, "field_type", None) == FieldType.reference]
    allowed_cache: dict[tuple[int, str, str | None], set[str]] = {}
    for row_idx, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        for sf in ref_sub_fields:
            cfg = getattr(sf, "config", None) or {}
            sid = cfg.get("reference_source_kpi_id")
            skey = cfg.get("reference_source_field_key")
            subkey = cfg.get("reference_source_sub_field_key")
            if not sid or not skey:
                continue
            cache_key = (int(sid), str(skey), str(subkey) if subkey else None)
            if cache_key not in allowed_cache:
                allowed = await get_reference_allowed_values(
                    db, int(sid), str(skey), org_id, source_sub_field_key=(str(subkey) if subkey else None)
                )
                allowed_cache[cache_key] = {_normalize_reference_value(a) for a in allowed}
            cell = item.get(sf.key)
            raw = cell if isinstance(cell, str) else str(cell) if cell is not None else ""
            normalized = _normalize_reference_value(raw)
            if _is_reference_empty_or_sentinel(normalized):
                # Normalize empty/NA-like values to None so downstream uses nulls.
                item[sf.key] = None
                continue
            if normalized not in allowed_cache[cache_key]:
                validation_errors.append(
                    {
                        "field_key": field.key,
                        "sub_field_key": sf.key,
                        "row_index": row_idx,
                        "value": raw,
                        "row": item,
                        "message": "Value does not exist in the referenced KPI field.",
                    }
                )
    multif_sub_fields = [
        s for s in (field.sub_fields or []) if getattr(s, "field_type", None) == FieldType.multi_reference
    ]
    multif_allowed_list: dict[tuple[int, str, str | None], list[str]] = {}
    for row_idx, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        for sf in multif_sub_fields:
            cfg = getattr(sf, "config", None) or {}
            sid = cfg.get("reference_source_kpi_id")
            skey = cfg.get("reference_source_field_key")
            subkey = cfg.get("reference_source_sub_field_key")
            if not sid or not skey:
                continue
            cache_key = (int(sid), str(skey), str(subkey) if subkey else None)
            if cache_key not in multif_allowed_list:
                multif_allowed_list[cache_key] = await get_reference_allowed_values(
                    db, int(sid), str(skey), org_id, source_sub_field_key=(str(subkey) if subkey else None)
                )
            allowed_list = multif_allowed_list[cache_key]
            allowed_norm = {_normalize_reference_value(a) for a in allowed_list}
            cell = item.get(sf.key)
            for tok in coerce_multi_reference_raw(cell):
                if isinstance(tok, dict):
                    s = None
                    for k in ("label", "text", "value", "name"):
                        if k in tok and tok[k] is not None:
                            s = str(tok[k])
                            break
                    if s is None:
                        continue
                else:
                    s = str(tok) if tok is not None else ""
                n = _normalize_reference_value(s)
                if _is_reference_empty_or_sentinel(n):
                    continue
                if n not in allowed_norm:
                    validation_errors.append(
                        {
                            "field_key": field.key,
                            "sub_field_key": sf.key,
                            "row_index": row_idx,
                            "value": str(cell),
                            "row": item,
                            "message": "One or more values do not exist in the referenced KPI field.",
                        }
                    )
                    break
            else:
                cleaned = filter_multi_reference_to_allowed(cell, allowed_list) if allowed_list else []
                item[sf.key] = cleaned if cleaned else None
    if validation_errors:
        raise EntryValidationError(validation_errors)

    items = [it for it in items if isinstance(it, dict) and not _is_multi_items_row_effectively_empty(it)]

    mode = _resolve_multi_items_import_mode(import_mode, append)
    if mode == "upsert":
        mk = (match_sub_field_key or "").strip()
        if not mk:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="match_sub_field_key is required when import_mode=upsert",
            )
        sub_by_key = {s.key: s for s in (field.sub_fields or [])}
        if mk not in sub_by_key:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="match_sub_field_key must be a defined sub-field key for this multi-line field",
            )
        match_ft = getattr(sub_by_key[mk], "field_type", None)
    else:
        mk = None
        match_ft = None

    existing_pairs = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=field)
    existing_list = [r for _, r in existing_pairs] if existing_pairs else []
    prev_count = len(existing_list)
    imported_count = len(items) if isinstance(items, list) else 0

    rows_updated = 0
    if mode == "append":
        new_rows = existing_list + items
        rows_added = imported_count
        rows_overridden = 0
    elif mode == "upsert":
        merged, rows_updated, rows_added = _upsert_merge_multi_line_items(
            existing_list, items, mk, match_ft
        )
        new_rows = merged
        rows_overridden = 0
    else:
        new_rows = items
        rows_added = imported_count
        rows_overridden = prev_count
    entry.user_id = current_user.id  # track last editor (optional)
    await _replace_multi_line_rows_from_dicts(db, entry_id=entry.id, field=field, rows=new_rows)
    await db.commit()

    return {
        "entry_id": entry.id,
        "field_id": field.id,
        "import_mode": mode,
        "append": mode == "append",
        "rows_added": rows_added,
        "rows_updated": rows_updated,
        "rows_overridden": rows_overridden,
        "match_sub_field_key": mk,
    }


@router.get("/multi-items/rows", response_model=MultiItemsListResponse)
async def list_multi_items_rows(
    entry_id: int = Query(...),
    field_id: int = Query(...),
    organization_id: int | None = Query(None),
    search: str | None = Query(None, description="Simple text search across row values"),
    sort_by: str | None = Query(None, description="Sub-field key to sort by"),
    sort_dir: str = Query("asc", pattern="^(asc|desc)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    filters: str | None = Query(
        None,
        description='Legacy: JSON object {"col_key":"substring"} (case-insensitive AND). '
        'Structured: {"_version":2,"conditions":[{"field":"col","op":"eq","value":"x"},{"logic":"and","field":"y","op":"gte","value":"10"}]} '
        '(ops: eq, neq, gt, gte, lt, lte, contains, not_contains, starts_with, ends_with).',
    ),
    editable_only: bool = Query(
        False,
        description="When true, only return rows where the current user can edit and/or delete the row.",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List multi_line_items rows for an entry+field with search, sort, and paging."""
    org_id = _org_id(current_user, organization_id)
    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entry not found")
    field = await _load_multi_items_field(db, org_id, field_id)
    if not field or field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")

    def _try_parse_iso_datetime(v: Any) -> datetime | None:
        if v is None:
            return None
        if isinstance(v, datetime):
            return v
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            try:
                # Accept YYYY-MM-DD or full ISO datetime
                return datetime.fromisoformat(s)
            except Exception:
                return None
        return None

    def _string_value_expr(cell) -> Any:
        # Used for text-like comparisons across typed columns.
        return func.coalesce(
            cast(cell.value_text, String()),
            cast(cell.value_json, String()),
            cast(cell.value_number, String()),
            cast(cell.value_boolean, String()),
            cast(cell.value_date, String()),
        )

    def _reference_like_exprs(cell) -> list[Any]:
        """
        Common expressions for reference / multi_reference stored either as plain text or as JSON object/array.
        We compare against:
        - value_text (when stored as label)
        - value_json->>'label' / ->>'value' / ->>'token' when stored as object
        - stringified JSON for array membership fallback (multi_reference)
        """
        return [
            cast(cell.value_text, String()),
            func.json_extract_path_text(cell.value_json, "label"),
            func.json_extract_path_text(cell.value_json, "value"),
            func.json_extract_path_text(cell.value_json, "token"),
            cast(cell.value_json, String()),
        ]

    def _build_sql_filter_clause(raw: Any) -> Any | None:
        """
        Build a SQLAlchemy boolean clause for filters payload when possible.
        Supports only structured _version=2 scalar filters (no reference_resolution).
        Returns None when payload is unsupported, so we can fall back to the existing slow path.
        """
        if not isinstance(raw, dict) or raw.get("_version") != 2:
            return None
        conds = raw.get("conditions")
        if not isinstance(conds, list) or not conds:
            return None

        # Map sub-field keys to (id, field_type)
        sf_map: dict[str, tuple[int, str]] = {}
        for sf in (field.sub_fields or []):
            k = str(getattr(sf, "key", "") or "")
            sid = getattr(sf, "id", None)
            if not k or sid is None:
                continue
            ft = getattr(getattr(sf, "field_type", None), "value", getattr(sf, "field_type", None))
            sf_map[k] = (int(sid), str(ft or ""))

        def exists_for(sub_field_id: int, pred) -> Any:
            return (
                select(func.count())
                .select_from(KpiMultiLineCell)
                .where(
                    and_(
                        KpiMultiLineCell.row_id == KpiMultiLineRow.id,
                        KpiMultiLineCell.sub_field_id == sub_field_id,
                        pred,
                    )
                )
                .correlate(KpiMultiLineRow)
                .scalar_subquery()
                > 0
            )

        expr = None
        for idx, c in enumerate(conds):
            if not isinstance(c, dict):
                return None
            if c.get("reference_resolution"):
                return None
            fk = str(c.get("field") or "").strip()
            op = str(c.get("op") or "").strip()
            if not fk or not op:
                return None
            if fk not in sf_map:
                return None
            sub_field_id, ft = sf_map[fk]

            values = c.get("values")
            value = c.get("value")
            use_values = isinstance(values, list) and len(values) > 0

            cell = KpiMultiLineCell
            clause_part = None

            if ft == "number":
                def to_num(x: Any) -> float | None:
                    try:
                        if x is None or (isinstance(x, str) and not x.strip()):
                            return None
                        return float(x)
                    except Exception:
                        return None
                if use_values and op in ("eq", "neq"):
                    nums = [to_num(x) for x in values]
                    nums = [n for n in nums if n is not None]
                    if not nums:
                        return None
                    pos = exists_for(sub_field_id, cell.value_number.in_(nums))
                    clause_part = pos if op == "eq" else ~pos
                else:
                    n = to_num(value)
                    if n is None:
                        return None
                    if op == "eq":
                        clause_part = exists_for(sub_field_id, cell.value_number == n)
                    elif op == "neq":
                        clause_part = ~exists_for(sub_field_id, cell.value_number == n)
                    elif op == "gt":
                        clause_part = exists_for(sub_field_id, cell.value_number > n)
                    elif op == "gte":
                        clause_part = exists_for(sub_field_id, cell.value_number >= n)
                    elif op == "lt":
                        clause_part = exists_for(sub_field_id, cell.value_number < n)
                    elif op == "lte":
                        clause_part = exists_for(sub_field_id, cell.value_number <= n)
                    else:
                        return None
            elif ft == "date":
                dt = _try_parse_iso_datetime(value)
                if dt is None:
                    return None
                if op == "eq":
                    clause_part = exists_for(sub_field_id, cell.value_date == dt)
                elif op == "neq":
                    clause_part = ~exists_for(sub_field_id, cell.value_date == dt)
                elif op == "gt":
                    clause_part = exists_for(sub_field_id, cell.value_date > dt)
                elif op == "gte":
                    clause_part = exists_for(sub_field_id, cell.value_date >= dt)
                elif op == "lt":
                    clause_part = exists_for(sub_field_id, cell.value_date < dt)
                elif op == "lte":
                    clause_part = exists_for(sub_field_id, cell.value_date <= dt)
                else:
                    return None
            elif ft == "boolean":
                if isinstance(value, bool):
                    b = value
                elif isinstance(value, str):
                    vs = value.strip().lower()
                    if vs == "true":
                        b = True
                    elif vs == "false":
                        b = False
                    else:
                        return None
                else:
                    return None
                if op == "eq":
                    clause_part = exists_for(sub_field_id, cell.value_boolean == b)
                elif op == "neq":
                    clause_part = ~exists_for(sub_field_id, cell.value_boolean == b)
                else:
                    return None
            else:
                # Treat as text-like.
                if use_values and op in ("eq", "neq"):
                    vals = [str(x).strip() for x in values if x is not None and str(x).strip()]
                    if not vals:
                        return None
                    if ft in ("reference", "multi_reference"):
                        # membership match for multi_reference arrays; exact/label match for reference objects
                        candidates = _reference_like_exprs(cell)
                        pred = or_(*[e.in_(vals) for e in candidates if e is not None])
                        pos = exists_for(sub_field_id, pred)
                    else:
                        pos = exists_for(sub_field_id, _string_value_expr(cell).in_(vals))
                    clause_part = pos if op == "eq" else ~pos
                else:
                    if value is None:
                        return None
                    s = str(value).strip()
                    if s == "":
                        return None
                    if ft in ("reference", "multi_reference"):
                        # For reference values we attempt exact match on text/label/value/token, and for multi_reference
                        # also allow JSON array membership via string contains on `"value"`.
                        exprs = _reference_like_exprs(cell)
                        if op == "eq":
                            pred = or_(*[(e == s) for e in exprs if e is not None])
                            # array membership fallback
                            pred = or_(pred, cast(cell.value_json, String()).ilike(f"%\\\"{s}\\\"%"))
                            clause_part = exists_for(sub_field_id, pred)
                        elif op == "neq":
                            pred = or_(*[(e == s) for e in exprs if e is not None])
                            pred = or_(pred, cast(cell.value_json, String()).ilike(f"%\\\"{s}\\\"%"))
                            clause_part = ~exists_for(sub_field_id, pred)
                        elif op == "contains":
                            pred = or_(*[(e.ilike(f"%{s}%")) for e in exprs if e is not None])
                            clause_part = exists_for(sub_field_id, pred)
                        elif op == "not_contains":
                            pred = or_(*[(e.ilike(f"%{s}%")) for e in exprs if e is not None])
                            clause_part = ~exists_for(sub_field_id, pred)
                        elif op == "starts_with":
                            pred = or_(*[(e.ilike(f"{s}%")) for e in exprs if e is not None])
                            clause_part = exists_for(sub_field_id, pred)
                        elif op == "ends_with":
                            pred = or_(*[(e.ilike(f"%{s}")) for e in exprs if e is not None])
                            clause_part = exists_for(sub_field_id, pred)
                        else:
                            return None
                        # done
                        logic = str(c.get("logic") or "").strip().lower()
                        if idx == 0:
                            expr = clause_part
                        else:
                            expr = or_(expr, clause_part) if logic == "or" else and_(expr, clause_part)
                        continue

                    vexpr = _string_value_expr(cell)
                    if op == "eq":
                        clause_part = exists_for(sub_field_id, vexpr == s)
                    elif op == "neq":
                        clause_part = ~exists_for(sub_field_id, vexpr == s)
                    elif op == "contains":
                        clause_part = exists_for(sub_field_id, vexpr.ilike(f"%{s}%"))
                    elif op == "not_contains":
                        clause_part = ~exists_for(sub_field_id, vexpr.ilike(f"%{s}%"))
                    elif op == "starts_with":
                        clause_part = exists_for(sub_field_id, vexpr.ilike(f"{s}%"))
                    elif op == "ends_with":
                        clause_part = exists_for(sub_field_id, vexpr.ilike(f"%{s}"))
                    else:
                        return None

            logic = str(c.get("logic") or "").strip().lower()
            if idx == 0:
                expr = clause_part
            else:
                expr = or_(expr, clause_part) if logic == "or" else and_(expr, clause_part)

        return expr

    raw_filters: Any | None = None
    sql_filters_clause = None
    if filters and not search:
        try:
            raw_filters = json.loads(filters)
            sql_filters_clause = _build_sql_filter_clause(raw_filters)
        except json.JSONDecodeError:
            raw_filters = None
            sql_filters_clause = None

    # Fast path: if caller isn't using search and filters are SQL-able, do true SQL pagination.
    # If filters are present but unsupported (e.g. reference_resolution), fall back to the existing slow path.
    use_fast_sql_paging = not search and (not filters or sql_filters_clause is not None)
    rows: list[tuple[int, dict]] = []

    # Restrict to sub_fields (columns) the user can view.
    # Important: don't call permission helpers per column (slow and can time out).
    viewable_keys: set[str] = set()
    sub_fields_payload: list[dict] = []
    is_org_admin = current_user.role.value in ("ORG_ADMIN", "SUPER_ADMIN")

    access_map = None
    can_view_kpi = False
    if not is_org_admin:
        # One shot: compute user's effective field access for this KPI.
        access_map = await get_user_field_access_for_kpi(db, current_user.id, entry.kpi_id)
        if access_map is None:
            # No explicit field-level access rows: rely on KPI-level visibility.
            can_view_kpi = await user_can_view_kpi(db, current_user.id, entry.kpi_id, org_id)

    for sf in (field.sub_fields or []):
        sf_key = getattr(sf, "key", "")
        sf_id = getattr(sf, "id", None)

        can_view = False
        if is_org_admin:
            can_view = True
        elif access_map is None:
            can_view = can_view_kpi
        else:
            perm = access_map.get((field.id, sf_id)) or access_map.get((field.id, None))
            can_view = perm in ("view", "data_entry")

        if can_view:
            viewable_keys.add(sf_key)
            ft = getattr(sf, "field_type", None)
            sub_fields_payload.append(
                {
                    "key": sf_key,
                    "name": getattr(sf, "name", sf_key),
                    # `field_type` should be an Enum(FieldType), but some datasets/migrations may load it as a plain string.
                    "field_type": ft.value if hasattr(ft, "value") else ft,
                    "is_required": getattr(sf, "is_required", False),
                }
            )
    if not use_fast_sql_paging:
        # Slow path only: we will load all rows into memory and need to restrict keys for search/sort/filter.
        # Keep all defined sub-field keys for search / sort / filters.
        subfield_key_set = {str(getattr(s, "key", "")) for s in (field.sub_fields or []) if getattr(s, "key", None)}
        rows = [
            (i, {k: v for k, v in r.items() if k in subfield_key_set})
            for i, r in rows
        ]

    # Filter by search
    if search:
        q = search.lower().strip()
        def matches(row: dict) -> bool:
            for v in row.values():
                if v is None:
                    continue
                s = str(v).lower()
                if q in s:
                    return True
            return False
        rows = [(i, r) for i, r in rows if matches(r)]

    # Advanced column filters (legacy substring map or structured _version 2)
    if filters:
        try:
            raw_filters = raw_filters if raw_filters is not None else json.loads(filters)
            if isinstance(raw_filters, dict):
                resolution_maps = None
                reference_field_types: dict[str, str] = {}
                for sf in field.sub_fields or []:
                    k = getattr(sf, "key", "")
                    ft = getattr(sf.field_type, "value", sf.field_type)
                    reference_field_types[str(k)] = str(ft)
                if raw_filters.get("_version") == 2:
                    conds = raw_filters.get("conditions")
                    if isinstance(conds, list):
                        # Only build reference resolution maps if filters actually require reference lookups.
                        # This can be very expensive; avoid it for pure scalar conditions (e.g. eq/contains on text/number/date).
                        needs_ref = False
                        for c in conds:
                            if not isinstance(c, dict):
                                continue
                            if c.get("reference_resolution"):
                                needs_ref = True
                                break
                            fk = c.get("field")
                            if fk is not None and reference_field_types.get(str(fk)) in ("reference", "multi_reference"):
                                needs_ref = True
                                break
                        if needs_ref:
                            resolution_maps = await build_reference_resolution_map(
                                db,
                                org_id,
                                entry.year,
                                field,
                                conds,
                                [r for _, r in rows],
                            )
                    rows = [
                        (i, r)
                        for i, r in rows
                        if row_passes_filters(
                            r,
                            raw_filters,
                            resolution_maps=resolution_maps,
                            reference_field_types=reference_field_types,
                        )
                    ]
                else:
                    rows = [(i, r) for i, r in rows if row_passes_filters(r, raw_filters)]
        except json.JSONDecodeError:
            # Ignore invalid filters payload
            pass

    # Sort (by row data; original index is preserved in the tuple)
    if sort_by:
        reverse = sort_dir == "desc"
        def sort_key(row: dict):
            v = row.get(sort_by)
            # Try numeric, then string
            try:
                return float(v)
            except (TypeError, ValueError):
                return str(v) if v is not None else ""
        try:
            rows = sorted(rows, key=lambda ir: sort_key(ir[1]), reverse=reverse)
        except Exception:
            pass

    # Per-row permissions for current user (edit/delete) - used for editable_only filtering and row payload.
    out_rows: list[MultiItemsRow] = []

    field_row_access_enabled = bool(getattr(field, "row_level_user_access_enabled", False))
    row_rule_map: dict[int, tuple[bool, bool]] = {}

    is_org_admin = current_user.role.value in ("ORG_ADMIN", "SUPER_ADMIN")

    # Common permissions when row-level access is disabled.
    # When row-level access is enabled, non-admins must be readonly unless explicitly allowed per-row.
    can_edit_common = await user_can_edit_multi_line_field(db, current_user.id, entry.kpi_id, field)
    can_delete_common = await user_can_edit_field(db, current_user.id, entry.kpi_id, field.id, None)

    if field_row_access_enabled:
        row_rules_res = await db.execute(
            select(
                KpiMultiLineRowAccess.row_index,
                KpiMultiLineRowAccess.can_edit,
                KpiMultiLineRowAccess.can_delete,
            ).where(
                KpiMultiLineRowAccess.user_id == current_user.id,
                KpiMultiLineRowAccess.entry_id == entry.id,
                KpiMultiLineRowAccess.field_id == field.id,
            )
        )
        row_rules = row_rules_res.all()
        for row_index, can_edit, can_delete in row_rules:
            row_rule_map[int(row_index)] = (bool(can_edit), bool(can_delete))

    # Visibility: when row-level access is enabled, non-admin users should only see rows explicitly assigned to them.
    # An access record implies view permission; can_edit/can_delete control actions.
    if field_row_access_enabled and not is_org_admin:
        pass

    # Optional: filter down to only rows the user can edit/delete.
    if editable_only:
        if not (field_row_access_enabled and not is_org_admin):
            # Row-level access disabled (or org/super admin): permissions are identical for all rows.
            if not (can_edit_common or can_delete_common):
                use_fast_sql_paging = True  # will return empty below

    if use_fast_sql_paging:
        # Determine which row_indices are visible to the user (row-level access) and/or editable_only.
        base_rows_stmt = select(KpiMultiLineRow.id, KpiMultiLineRow.row_index).where(
            KpiMultiLineRow.entry_id == entry.id,
            KpiMultiLineRow.field_id == field.id,
        )
        if sql_filters_clause is not None:
            base_rows_stmt = base_rows_stmt.where(sql_filters_clause)
        if field_row_access_enabled and not is_org_admin:
            # Only rows with an access record for this user are visible.
            base_rows_stmt = (
                base_rows_stmt.join(
                    KpiMultiLineRowAccess,
                    and_(
                        KpiMultiLineRowAccess.entry_id == KpiMultiLineRow.entry_id,
                        KpiMultiLineRowAccess.field_id == KpiMultiLineRow.field_id,
                        KpiMultiLineRowAccess.row_index == KpiMultiLineRow.row_index,
                    ),
                )
                .where(KpiMultiLineRowAccess.user_id == current_user.id)
            )
            if editable_only:
                base_rows_stmt = base_rows_stmt.where(
                    (KpiMultiLineRowAccess.can_edit == True) | (KpiMultiLineRowAccess.can_delete == True)  # noqa: E712
                )
        elif editable_only and not (can_edit_common or can_delete_common):
            # No row is editable for this user.
            total = 0
            return MultiItemsListResponse(
                total=0,
                page=page,
                page_size=page_size,
                rows=[],
                sub_fields=sub_fields_payload,
            )

        total = int(
            (
                await db.execute(
                    select(func.count()).select_from(base_rows_stmt.subquery())
                )
            ).scalar_one()
            or 0
        )
        start = (page - 1) * page_size

        order_stmt = base_rows_stmt
        if sort_by:
            # SQL sort on a specific sub-field to avoid the slow in-memory path on large datasets.
            # This keeps paging in SQL even for 20k+ rows.
            sort_sf = next((s for s in (field.sub_fields or []) if str(getattr(s, "key", "")) == str(sort_by)), None)
            sort_sf_id = int(getattr(sort_sf, "id", 0) or 0) if sort_sf is not None else 0
            sort_ft = getattr(getattr(sort_sf, "field_type", None), "value", getattr(sort_sf, "field_type", None)) if sort_sf is not None else None
            if sort_sf_id > 0:
                order_cell = KpiMultiLineCell
                order_stmt = order_stmt.outerjoin(
                    order_cell,
                    and_(
                        order_cell.row_id == KpiMultiLineRow.id,
                        order_cell.sub_field_id == sort_sf_id,
                    ),
                )
                if str(sort_ft) == "number":
                    sort_expr = order_cell.value_number
                elif str(sort_ft) == "date":
                    sort_expr = order_cell.value_date
                elif str(sort_ft) == "boolean":
                    # bool sorts false < true; keep nulls last
                    sort_expr = order_cell.value_boolean
                else:
                    # text / json / reference-like: sort by stringified value
                    sort_expr = func.coalesce(
                        cast(order_cell.value_text, String()),
                        cast(order_cell.value_json, String()),
                        cast(order_cell.value_number, String()),
                        cast(order_cell.value_boolean, String()),
                        cast(order_cell.value_date, String()),
                    )
                reverse = sort_dir == "desc"
                order_stmt = order_stmt.order_by(nulls_last(sort_expr.desc() if reverse else sort_expr.asc()))
            else:
                order_stmt = order_stmt.order_by(KpiMultiLineRow.row_index)
        else:
            order_stmt = order_stmt.order_by(KpiMultiLineRow.row_index)

        paged_rows_res = await db.execute(order_stmt.offset(start).limit(page_size))
        page_rows = list(paged_rows_res.all())  # [(id, row_index)]
        if not page_rows:
            return MultiItemsListResponse(
                total=total,
                page=page,
                page_size=page_size,
                rows=[],
                sub_fields=sub_fields_payload,
            )
        row_ids = [int(r[0]) for r in page_rows]
        row_index_by_id = {int(rid): int(ridx) for rid, ridx in page_rows}

        cell_res = await db.execute(
            select(
                KpiMultiLineCell.row_id,
                KPIFieldSubField.key,
                KpiMultiLineCell.value_text,
                KpiMultiLineCell.value_number,
                KpiMultiLineCell.value_boolean,
                KpiMultiLineCell.value_date,
                KpiMultiLineCell.value_json,
            )
            .select_from(KpiMultiLineCell)
            .join(KPIFieldSubField, KPIFieldSubField.id == KpiMultiLineCell.sub_field_id)
            .where(KpiMultiLineCell.row_id.in_(row_ids))
        )

        row_data_by_index: dict[int, dict[str, Any]] = {row_index_by_id[rid]: {} for rid in row_ids}
        for row_id, key, vt, vn, vb, vd, vj in cell_res.all():
            idx = row_index_by_id.get(int(row_id))
            if idx is None or not key:
                continue
            # Reuse helper behavior from _cell_value_raw (inline for speed)
            if vj is not None:
                raw = vj
            elif vt is not None:
                raw = vt
            elif vn is not None:
                raw = vn
            elif vb is not None:
                raw = vb
            elif vd is not None:
                raw = vd.isoformat() if hasattr(vd, "isoformat") else str(vd)
            else:
                raw = None
            row_data_by_index[idx][str(key)] = raw

        for rid, row_index in page_rows:
            orig_index = int(row_index)
            r = row_data_by_index.get(orig_index, {})
            if field_row_access_enabled and not is_org_admin:
                rule = row_rule_map.get(int(orig_index))
                can_edit = rule[0] if rule else False
                can_delete = rule[1] if rule else False
            else:
                can_edit = can_edit_common
                can_delete = can_delete_common
            r_visible = {k: v for k, v in (r or {}).items() if k in viewable_keys}
            out_rows.append(MultiItemsRow(index=orig_index, data=r_visible, can_edit=can_edit, can_delete=can_delete))

        return MultiItemsListResponse(
            total=total,
            page=page,
            page_size=page_size,
            rows=out_rows,
            sub_fields=sub_fields_payload,
        )

    # Slow path: load all rows into memory for complex search/filters/sort_by.
    # (This is kept for feature parity; we can iterate to push these into SQL later.)
    rows: list[tuple[int, dict]] = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=field)

    # Restrict keys for search/sort/filter operations.
    subfield_key_set = {str(getattr(s, "key", "")) for s in (field.sub_fields or []) if getattr(s, "key", None)}
    rows = [(i, {k: v for k, v in r.items() if k in subfield_key_set}) for i, r in rows]

    # Filter by search
    if search:
        q = search.lower().strip()

        def matches(row: dict) -> bool:
            for v in row.values():
                if v is None:
                    continue
                s = str(v).lower()
                if q in s:
                    return True
            return False

        rows = [(i, r) for i, r in rows if matches(r)]

    # Advanced column filters (legacy substring map or structured _version 2)
    if filters:
        try:
            raw_filters = json.loads(filters)
            if isinstance(raw_filters, dict):
                resolution_maps = None
                reference_field_types: dict[str, str] = {}
                for sf in field.sub_fields or []:
                    k = getattr(sf, "key", "")
                    ft = getattr(sf.field_type, "value", sf.field_type)
                    reference_field_types[str(k)] = str(ft)
                if raw_filters.get("_version") == 2:
                    conds = raw_filters.get("conditions")
                    if isinstance(conds, list):
                        needs_ref = False
                        for c in conds:
                            if not isinstance(c, dict):
                                continue
                            if c.get("reference_resolution"):
                                needs_ref = True
                                break
                            fk = c.get("field")
                            if fk is not None and reference_field_types.get(str(fk)) in ("reference", "multi_reference"):
                                needs_ref = True
                                break
                        if needs_ref:
                            resolution_maps = await build_reference_resolution_map(
                                db,
                                org_id,
                                entry.year,
                                field,
                                conds,
                                [r for _, r in rows],
                            )
                    rows = [
                        (i, r)
                        for i, r in rows
                        if row_passes_filters(
                            r,
                            raw_filters,
                            resolution_maps=resolution_maps,
                            reference_field_types=reference_field_types,
                        )
                    ]
                else:
                    rows = [(i, r) for i, r in rows if row_passes_filters(r, raw_filters)]
        except json.JSONDecodeError:
            pass

    # Sort (by row data; original index is preserved in the tuple)
    if sort_by:
        reverse = sort_dir == "desc"

        def sort_key(row: dict):
            v = row.get(sort_by)
            try:
                return float(v)
            except (TypeError, ValueError):
                return str(v) if v is not None else ""

        try:
            rows = sorted(rows, key=lambda ir: sort_key(ir[1]), reverse=reverse)
        except Exception:
            pass

    # Visibility: when row-level access is enabled, non-admin users should only see rows explicitly assigned to them.
    if field_row_access_enabled and not is_org_admin:
        rows = [(i, r) for (i, r) in rows if int(i) in row_rule_map]

    # Optional: filter down to only rows the user can edit/delete.
    if editable_only:
        if field_row_access_enabled and not is_org_admin:

            def row_can_edit_or_delete(orig_index: int) -> bool:
                rule = row_rule_map.get(int(orig_index))
                return bool(rule and (rule[0] or rule[1]))

            rows = [(i, r) for (i, r) in rows if row_can_edit_or_delete(i)]
        else:
            if not (can_edit_common or can_delete_common):
                rows = []

    total = len(rows)
    start = (page - 1) * page_size
    end = start + page_size
    paged_rows = rows[start:end]

    for orig_index, r in paged_rows:
        if field_row_access_enabled and not is_org_admin:
            # Non-admins: readonly unless explicitly allowed for this row.
            rule = row_rule_map.get(int(orig_index))
            can_edit = rule[0] if rule else False
            can_delete = rule[1] if rule else False
        else:
            # Row-level disabled, or org/super admin: use normal permissions.
            can_edit = can_edit_common
            can_delete = can_delete_common

        r_visible = {k: v for k, v in r.items() if k in viewable_keys}
        out_rows.append(MultiItemsRow(index=orig_index, data=r_visible, can_edit=can_edit, can_delete=can_delete))

    return MultiItemsListResponse(
        total=total,
        page=page,
        page_size=page_size,
        rows=out_rows,
        sub_fields=sub_fields_payload,
    )


@router.post("/multi-items/rows", response_model=MultiItemsRow)
async def add_multi_items_row(
    entry_id: int = Query(...),
    field_id: int = Query(...),
    row: dict = Body(...),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Append a new row to relational multi_line_items storage for an entry+field."""
    org_id = _org_id(current_user, organization_id)
    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry or entry.is_locked:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Entry not editable")
    field = await _load_multi_items_field(db, org_id, field_id)
    if not field or field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")
    can_add = await user_can_add_row_multi_line_field(db, current_user.id, field.kpi_id, field.id)
    if not can_add:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to add rows to this field")

    # Normalize special sub-field types (e.g. mixed_list) for consistent storage.
    normalized_row = row if isinstance(row, dict) else {}
    key_to_sub = {getattr(s, "key", None): s for s in (field.sub_fields or []) if getattr(s, "key", None)}
    for k, v in list(normalized_row.items()):
        sub = key_to_sub.get(k)
        if not sub:
            continue
        if getattr(sub, "field_type", None) == FieldType.mixed_list:
            normalized_row[k] = coerce_mixed_list_raw(v) or None

    max_idx_res = await db.execute(
        select(func.max(KpiMultiLineRow.row_index)).where(
            KpiMultiLineRow.entry_id == entry.id,
            KpiMultiLineRow.field_id == field.id,
        )
    )
    max_idx = max_idx_res.scalar_one_or_none()
    new_index = int(max_idx) + 1 if max_idx is not None else 0

    mlr = KpiMultiLineRow(entry_id=entry.id, field_id=field.id, row_index=new_index)
    db.add(mlr)
    await db.flush()  # populate mlr.id

    def _add_cell(sub, raw_val: Any) -> None:
        c = KpiMultiLineCell(row_id=mlr.id, sub_field_id=int(getattr(sub, "id")))
        ft = getattr(sub, "field_type", None)
        ft_s = ft.value if hasattr(ft, "value") else str(ft)
        if raw_val is None:
            pass
        elif ft_s == "number":
            try:
                c.value_number = float(raw_val)
            except Exception:
                c.value_text = str(raw_val)
        elif ft_s == "boolean":
            if isinstance(raw_val, bool):
                c.value_boolean = raw_val
            else:
                s = str(raw_val).strip().lower()
                if s in ("true", "yes", "1"):
                    c.value_boolean = True
                elif s in ("false", "no", "0"):
                    c.value_boolean = False
                else:
                    c.value_text = str(raw_val)
        elif ft_s == "date":
            # UI often stores ISO date string; keep it as text for consistency.
            c.value_text = str(raw_val)
        elif ft_s in ("reference", "multi_reference", "mixed_list", "attachment"):
            if isinstance(raw_val, (dict, list)):
                c.value_json = raw_val
            else:
                c.value_text = str(raw_val)
        else:
            if isinstance(raw_val, (dict, list)):
                c.value_json = raw_val
            else:
                c.value_text = str(raw_val)
        db.add(c)

    for k, v in normalized_row.items():
        sub = key_to_sub.get(k)
        if not sub:
            continue
        _add_cell(sub, v)

    await db.commit()
    return MultiItemsRow(index=new_index, data=normalized_row)


@router.put("/multi-items/rows/{row_index}", response_model=MultiItemsRow)
async def update_multi_items_row(
    row_index: int,
    entry_id: int = Query(...),
    field_id: int = Query(...),
    row: dict = Body(...),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update a single row in relational multi_line_items storage."""
    org_id = _org_id(current_user, organization_id)
    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry or entry.is_locked:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Entry not editable")
    field = await _load_multi_items_field(db, org_id, field_id)
    if not field or field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")
    can_edit = await user_can_edit_row(db, current_user.id, entry.id, field.id, row_index)
    if not can_edit:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to edit this row")

    mlr_res = await db.execute(
        select(KpiMultiLineRow)
        .where(
            KpiMultiLineRow.entry_id == entry.id,
            KpiMultiLineRow.field_id == field.id,
            KpiMultiLineRow.row_index == row_index,
        )
        .options(selectinload(KpiMultiLineRow.cells))
    )
    mlr = mlr_res.scalar_one_or_none()
    if not mlr:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Row index out of range")

    existing_cells = {int(c.sub_field_id): c for c in (mlr.cells or []) if getattr(c, "sub_field_id", None) is not None}
    key_to_sub = {getattr(s, "key", None): s for s in (field.sub_fields or []) if getattr(s, "key", None)}

    def _set_cell_value(cell: KpiMultiLineCell, sub: Any, raw_val: Any) -> None:
        cell.value_text = None
        cell.value_number = None
        cell.value_json = None
        cell.value_boolean = None
        cell.value_date = None
        ft = getattr(sub, "field_type", None)
        ft_s = ft.value if hasattr(ft, "value") else str(ft)
        if raw_val is None:
            return
        if ft_s == "number":
            try:
                cell.value_number = float(raw_val)
            except Exception:
                cell.value_text = str(raw_val)
            return
        if ft_s == "boolean":
            if isinstance(raw_val, bool):
                cell.value_boolean = raw_val
            else:
                s = str(raw_val).strip().lower()
                if s in ("true", "yes", "1"):
                    cell.value_boolean = True
                elif s in ("false", "no", "0"):
                    cell.value_boolean = False
                else:
                    cell.value_text = str(raw_val)
            return
        if ft_s == "date":
            cell.value_text = str(raw_val)
            return
        if ft_s in ("reference", "multi_reference", "mixed_list", "attachment"):
            if isinstance(raw_val, (dict, list)):
                cell.value_json = raw_val
            else:
                cell.value_text = str(raw_val)
            return
        if isinstance(raw_val, (dict, list)):
            cell.value_json = raw_val
        else:
            cell.value_text = str(raw_val)

    # Merge row: only update cells for sub_fields the user can edit
    for col_key, col_value in (row or {}).items():
        sub = key_to_sub.get(col_key)
        if sub is None:
            continue
        if not await user_can_edit_field(db, current_user.id, entry.kpi_id, field.id, getattr(sub, "id", None)):
            continue
        next_val = (coerce_mixed_list_raw(col_value) or None) if getattr(sub, "field_type", None) == FieldType.mixed_list else col_value
        sub_id = int(getattr(sub, "id"))
        cell = existing_cells.get(sub_id)
        if cell is None:
            cell = KpiMultiLineCell(row_id=mlr.id, sub_field_id=sub_id)
            db.add(cell)
            existing_cells[sub_id] = cell
        _set_cell_value(cell, sub, next_val)

    await db.commit()
    # Return row in legacy dict shape
    rows = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=field, row_indices=[row_index])
    data = rows[0][1] if rows else {}
    return MultiItemsRow(index=row_index, data=data)


@router.post("/multi-items/rows/{row_index}/cell", response_model=MultiItemsRow)
async def update_multi_items_row_cell(
    row_index: int,
    entry_id: int = Query(...),
    field_id: int = Query(...),
    key: str = Query(..., description="Sub-field key to update"),
    value: str | None = Body(None, description="New value for the sub-field (stringified)"),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Update a single cell in a multi_line_items row.

    Intended for cases like per-row file upload where we only want to change one sub-field
    without overwriting the rest of the row.
    """
    org_id = _org_id(current_user, organization_id)
    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry or entry.is_locked:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Entry not editable")
    field = await _load_multi_items_field(db, org_id, field_id)
    if not field or field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")
    can_edit = await user_can_edit_row(db, current_user.id, entry.id, field.id, row_index)
    if not can_edit:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to edit this row")
    sub = next((s for s in (field.sub_fields or []) if getattr(s, "key", None) == key), None)
    if not sub:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sub-field not found")
    can_edit_cell = await user_can_edit_field(db, current_user.id, entry.kpi_id, field.id, getattr(sub, "id", None))
    if not can_edit_cell:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to edit this column")

    mlr_res = await db.execute(
        select(KpiMultiLineRow)
        .where(
            KpiMultiLineRow.entry_id == entry.id,
            KpiMultiLineRow.field_id == field.id,
            KpiMultiLineRow.row_index == row_index,
        )
        .options(selectinload(KpiMultiLineRow.cells))
    )
    mlr = mlr_res.scalar_one_or_none()
    if not mlr:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Row index out of range")

    sub_id = int(getattr(sub, "id"))
    cell = next((c for c in (mlr.cells or []) if int(getattr(c, "sub_field_id", -1)) == sub_id), None)
    if cell is None:
        cell = KpiMultiLineCell(row_id=mlr.id, sub_field_id=sub_id)
        db.add(cell)

    raw_val: Any = value
    if getattr(sub, "field_type", None) == FieldType.mixed_list:
        raw_val = coerce_mixed_list_raw(value) or None

    # Typed set (similar to update_multi_items_row)
    cell.value_text = None
    cell.value_number = None
    cell.value_json = None
    cell.value_boolean = None
    cell.value_date = None
    ft = getattr(sub, "field_type", None)
    ft_s = ft.value if hasattr(ft, "value") else str(ft)
    if raw_val is None:
        pass
    elif ft_s == "number":
        try:
            cell.value_number = float(raw_val)
        except Exception:
            cell.value_text = str(raw_val)
    elif ft_s == "boolean":
        if isinstance(raw_val, bool):
            cell.value_boolean = raw_val
        else:
            s = str(raw_val).strip().lower()
            if s in ("true", "yes", "1"):
                cell.value_boolean = True
            elif s in ("false", "no", "0"):
                cell.value_boolean = False
            else:
                cell.value_text = str(raw_val)
    elif ft_s == "date":
        cell.value_text = str(raw_val)
    elif ft_s in ("reference", "multi_reference", "mixed_list", "attachment"):
        if isinstance(raw_val, (dict, list)):
            cell.value_json = raw_val
        else:
            cell.value_text = str(raw_val)
    else:
        if isinstance(raw_val, (dict, list)):
            cell.value_json = raw_val
        else:
            cell.value_text = str(raw_val)

    await db.commit()
    rows = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=field, row_indices=[row_index])
    data = rows[0][1] if rows else {}
    return MultiItemsRow(index=row_index, data=data)


@router.delete("/multi-items/rows/{row_index}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_multi_items_row(
    row_index: int,
    entry_id: int = Query(...),
    field_id: int = Query(...),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a single row in relational multi_line_items storage."""
    org_id = _org_id(current_user, organization_id)
    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry or entry.is_locked:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Entry not editable")
    field = await _load_multi_items_field(db, org_id, field_id)
    if not field or field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")
    can_delete = await user_can_delete_row(db, current_user.id, entry.id, field.id, row_index)
    if not can_delete:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to delete this row")

    mlr_res = await db.execute(
        select(KpiMultiLineRow).where(
            KpiMultiLineRow.entry_id == entry.id,
            KpiMultiLineRow.field_id == field.id,
            KpiMultiLineRow.row_index == row_index,
        )
    )
    mlr = mlr_res.scalar_one_or_none()
    if not mlr:
        return
    await db.delete(mlr)
    await db.flush()
    await _reindex_multi_line_rows(db, entry_id=entry.id, field_id=field.id)
    await _reindex_row_access_after_delete(
        db,
        entry_id=entry.id,
        field_id=field.id,
        deleted_indices={row_index},
    )
    await db.commit()


@router.post("/multi-items/rows/bulk-delete", status_code=status.HTTP_204_NO_CONTENT)
async def bulk_delete_multi_items_rows(
    entry_id: int = Query(...),
    field_id: int = Query(...),
    indices: list[int] = Body(..., embed=True),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Bulk delete rows in relational multi_line_items storage by index list."""
    org_id = _org_id(current_user, organization_id)
    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry or entry.is_locked:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Entry not editable")
    field = await _load_multi_items_field(db, org_id, field_id)
    if not field or field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")

    raw_index_set = {i for i in indices if isinstance(i, int) and i >= 0}
    if not raw_index_set:
        return
    # Only delete rows the user is allowed to delete (record-level or field-level)
    index_set: set[int] = set()
    for idx in raw_index_set:
        if await user_can_delete_row(db, current_user.id, entry.id, field.id, idx):
            index_set.add(int(idx))
    if not index_set:
        return

    # Delete rows matching indices
    del_res = await db.execute(
        select(KpiMultiLineRow).where(
            KpiMultiLineRow.entry_id == entry.id,
            KpiMultiLineRow.field_id == field.id,
            KpiMultiLineRow.row_index.in_(sorted(index_set)),
        )
    )
    for r in list(del_res.scalars().all()):
        await db.delete(r)
    await db.flush()

    await _reindex_multi_line_rows(db, entry_id=entry.id, field_id=field.id)
    await _reindex_row_access_after_delete(
        db,
        entry_id=entry.id,
        field_id=field.id,
        deleted_indices=index_set,
    )
    await db.commit()


@router.post("/multi-items/sync-from-api")
async def sync_multi_items_from_api(
    entry_id: int = Query(...),
    field_id: int = Query(...),
    organization_id: int | None = Query(None),
    sync_mode: str = Query(
        "override",
        description="override = replace existing; append = append rows; upsert = update by match_sub_field_key or add",
        pattern="^(override|append|upsert)$",
    ),
    match_sub_field_key: str | None = Query(
        None,
        description="Required when sync_mode=upsert: sub-field key to match existing rows",
    ),
    api_url: str | None = Query(None, description="Optional override for field config multi_items_api_endpoint_url"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """
    Sync a multi_line_items field from its own API endpoint.

    - API URL is taken from KPIField.config['multi_items_api_endpoint_url'].
    - Request payload: { year, kpi_id, field_id, field_key, organization_id, entry_id }.
    - Expected response:
        {
          "year": 2026,
          "items": [ { ...row1 }, { ...row2 }, ... ]
        }
    - sync_mode (query param, from UI) always wins. Any override / append / merge hints in the API body
      (e.g. override_existing) are ignored.
    - sync_mode:
        - "override": replace existing rows with items
        - "append": append items to existing rows
        - "upsert": rows whose match field equals an existing row (normalized) are merged; others appended
    """
    org_id = _org_id(current_user, organization_id)
    # Load entry and field
    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry or entry.is_locked:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Entry not editable")

    field_res = await db.execute(
        select(KPIField)
        .join(KPI, KPI.id == KPIField.kpi_id)
        .where(KPIField.id == field_id, KPI.organization_id == org_id)
    )
    field = field_res.scalar_one_or_none()
    if not field or field.field_type != FieldType.multi_line_items or field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-line field not found")

    can_edit = await user_can_edit_kpi(db, current_user.id, field.kpi_id)
    if not can_edit:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to edit this KPI")

    cfg = getattr(field, "config", None) or {}
    configured_url = (cfg.get("multi_items_api_endpoint_url") or "").strip()
    final_api_url = (api_url or configured_url or "").strip()
    if not final_api_url:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No multi_items_api_endpoint_url configured for this field")

    payload = {
        "year": entry.year,
        "kpi_id": field.kpi_id,
        "field_id": field.id,
        "field_key": field.key,
        "organization_id": org_id,
        "entry_id": entry.id,
    }
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(final_api_url, json=payload)
            resp.raise_for_status()
            data = resp.json()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Field API error: {e!s}",
        )
    if not isinstance(data, dict):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Field API response must be a JSON object")
    resp_year = data.get("year")
    items = data.get("items")
    if resp_year is not None and int(resp_year) != int(entry.year):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Year in field API response does not match entry year")
    if not isinstance(items, list):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Field API response 'items' must be a list")

    # Load existing rows (relational)
    existing_pairs = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=field)
    existing_rows: list[dict] = [r for _, r in existing_pairs] if existing_pairs else []

    effective_mode = (sync_mode or "override").lower()
    item_dicts = [
        dict(x)
        for x in items
        if isinstance(x, dict) and not _is_multi_items_row_effectively_empty(dict(x))
    ]

    if effective_mode == "upsert":
        mk = (match_sub_field_key or "").strip()
        if not mk:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="match_sub_field_key is required when sync_mode=upsert",
            )
        sub_by_key = {s.key: s for s in (field.sub_fields or [])}
        if mk not in sub_by_key:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="match_sub_field_key must be a defined sub-field key for this multi-line field",
            )
        match_ft = getattr(sub_by_key[mk], "field_type", None)
        new_rows, rows_updated, rows_added = _upsert_merge_multi_line_items(
            existing_rows, item_dicts, mk, match_ft
        )
    elif effective_mode == "append":
        new_rows = existing_rows + item_dicts
        rows_updated = 0
        rows_added = len(item_dicts)
    else:
        new_rows = item_dicts
        rows_updated = 0
        rows_added = len(item_dicts)

    await _replace_multi_line_rows_from_dicts(db, entry_id=entry.id, field=field, rows=new_rows)
    await db.commit()
    out: dict = {
        "entry_id": entry.id,
        "field_id": field.id,
        "rows_imported": len(item_dicts),
    }
    if effective_mode == "upsert":
        out["rows_updated"] = rows_updated
        out["rows_appended"] = rows_added
    return out


@router.post("/multi-items/import-from-year")
async def import_multi_items_from_year(
    entry_id: int = Query(...),
    field_id: int = Query(...),
    source_year: int = Query(..., ge=1900, le=3000),
    source_period_key: str | None = Query(
        None,
        description="Optional: period_key of source entry. Defaults to target entry's period_key.",
    ),
    import_mode: str | None = Query(
        None,
        description="replace (default), append, or upsert (requires match_sub_field_key)",
        pattern="^(replace|append|upsert)$",
    ),
    match_sub_field_key: str | None = Query(
        None,
        description="Sub-field key used to match rows when import_mode=upsert (same normalized value => update row)",
    ),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import multi_line_items rows from a selected year's entry into this entry+field."""
    org_id = _org_id(current_user, organization_id)

    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entry not found")
    if entry.is_locked:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Entry is locked")

    field = await _load_multi_items_field(db, org_id, field_id)
    if not field or field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")

    can_add = await user_can_add_row_multi_line_field(db, current_user.id, field.kpi_id, field.id)
    if not can_add:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to add rows to this field")

    can_edit = await user_can_edit_multi_line_field(db, current_user.id, entry.kpi_id, field)
    if not can_edit:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to edit this field")

    src_period = (source_period_key if source_period_key is not None else (getattr(entry, "period_key", "") or "")).strip()
    src_entry_res = await db.execute(
        select(KPIEntry).where(
            KPIEntry.kpi_id == entry.kpi_id,
            KPIEntry.organization_id == org_id,
            KPIEntry.year == int(source_year),
            KPIEntry.period_key == src_period,
        )
    )
    src_entry = src_entry_res.scalar_one_or_none()
    if not src_entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Source entry not found for selected year/period")

    incoming_pairs = await _load_multi_line_row_dicts(db, entry_id=src_entry.id, field=field)
    incoming_items = [
        dict(x)
        for _, x in (incoming_pairs or [])
        if isinstance(x, dict) and not _is_multi_items_row_effectively_empty(dict(x))
    ]

    mode = _resolve_multi_items_import_mode(import_mode, append_legacy=False)
    if mode == "upsert":
        mk = (match_sub_field_key or "").strip()
        if not mk:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="match_sub_field_key is required when import_mode=upsert")
        sub_by_key = {s.key: s for s in (field.sub_fields or [])}
        if mk not in sub_by_key:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="match_sub_field_key must be a defined sub-field key for this multi-line field")
        match_ft = getattr(sub_by_key[mk], "field_type", None)
    else:
        mk = None
        match_ft = None

    existing_pairs = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=field)
    existing_list = [r for _, r in existing_pairs] if existing_pairs else []
    prev_count = len(existing_list)

    rows_updated = 0
    if mode == "append":
        new_rows = existing_list + incoming_items
        rows_added = len(incoming_items)
        rows_overridden = 0
    elif mode == "upsert":
        merged, rows_updated, rows_added = _upsert_merge_multi_line_items(existing_list, incoming_items, mk, match_ft)
        new_rows = merged
        rows_overridden = 0
    else:
        new_rows = incoming_items
        rows_added = len(incoming_items)
        rows_overridden = prev_count

    entry.user_id = current_user.id
    await _replace_multi_line_rows_from_dicts(db, entry_id=entry.id, field=field, rows=new_rows)
    await db.commit()

    out: dict = {
        "entry_id": entry.id,
        "field_id": field.id,
        "source_entry_id": src_entry.id,
        "source_year": int(source_year),
        "source_period_key": src_period,
        "import_mode": mode,
        "rows_imported": len(incoming_items),
    }
    if mode == "upsert":
        out["rows_updated"] = rows_updated
        out["rows_appended"] = rows_added
    else:
        out["rows_added"] = rows_added
        out["rows_overridden"] = rows_overridden
    return out


@router.get("/multi-items/available-source-years")
async def list_multi_items_available_source_years(
    kpi_id: int = Query(...),
    field_id: int = Query(...),
    target_year: int = Query(..., ge=1900, le=3000),
    period_key: str | None = Query(None, description="Optional: period_key to match. If omitted, any period_key is allowed."),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return previous years (< target_year) that have any stored rows for this multi_line_items field."""
    org_id = _org_id(current_user, organization_id)

    field = await _load_multi_items_field(db, org_id, field_id)
    if not field or field.kpi_id != int(kpi_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")

    # Any user who can view this KPI should be able to see available years.
    can_view = await user_can_view_kpi(db, current_user.id, int(kpi_id))
    if not can_view:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")

    stmt = (
        select(distinct(KPIEntry.year))
        .select_from(KPIEntry)
        .join(KpiMultiLineRow, KpiMultiLineRow.entry_id == KPIEntry.id)
        .where(
            KPIEntry.organization_id == org_id,
            KPIEntry.kpi_id == int(kpi_id),
            KPIEntry.year < int(target_year),
            KpiMultiLineRow.field_id == int(field_id),
        )
    )
    if period_key is not None:
        stmt = stmt.where(KPIEntry.period_key == str(period_key))
    res = await db.execute(stmt)
    years = sorted({int(r[0]) for r in res.all() if r and r[0] is not None}, reverse=True)
    return {"years": years}


@router.get("/multi-items/export")
async def export_multi_items_csv(
    entry_id: int = Query(...),
    field_id: int = Query(...),
    organization_id: int | None = Query(None),
    search: str | None = Query(None),
    sort_by: str | None = Query(None),
    sort_dir: str = Query("asc", pattern="^(asc|desc)$"),
    filters: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Export multi_line_items rows for an entry+field as CSV, honoring search, sort, and filters.
    """
    org_id = _org_id(current_user, organization_id)
    entry_res = await db.execute(
        select(KPIEntry).where(KPIEntry.id == entry_id, KPIEntry.organization_id == org_id)
    )
    entry = entry_res.scalar_one_or_none()
    if not entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entry not found")
    field = await _load_multi_items_field(db, org_id, field_id)
    if not field or field.kpi_id != entry.kpi_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")

    pairs = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=field)
    rows: list[dict] = [r for _, r in pairs] if pairs else []

    # Filter by search
    if search:
        q = search.lower().strip()

        def matches(row: dict) -> bool:
            for v in row.values():
                if v is None:
                    continue
                s = str(v).lower()
                if q in s:
                    return True
            return False

        rows = [r for r in rows if isinstance(r, dict) and matches(r)]

    # Advanced filters (same payload as list_multi_items_rows)
    if filters:
        try:
            raw_filters = json.loads(filters)
            if isinstance(raw_filters, dict):
                resolution_maps = None
                reference_field_types: dict[str, str] = {}
                for sf in field.sub_fields or []:
                    k = getattr(sf, "key", "")
                    ft = getattr(sf.field_type, "value", sf.field_type)
                    reference_field_types[str(k)] = str(ft)
                if raw_filters.get("_version") == 2:
                    conds = raw_filters.get("conditions")
                    if isinstance(conds, list):
                        needs_ref = False
                        for c in conds:
                            if not isinstance(c, dict):
                                continue
                            if c.get("reference_resolution"):
                                needs_ref = True
                                break
                            fk = c.get("field")
                            if fk is not None and reference_field_types.get(str(fk)) in ("reference", "multi_reference"):
                                needs_ref = True
                                break
                        if needs_ref:
                            resolution_maps = await build_reference_resolution_map(
                                db,
                                org_id,
                                entry.year,
                                field,
                                conds,
                                [r for r in rows if isinstance(r, dict)],
                            )
                    rows = [
                        r
                        for r in rows
                        if isinstance(r, dict)
                        and row_passes_filters(
                            r,
                            raw_filters,
                            resolution_maps=resolution_maps,
                            reference_field_types=reference_field_types,
                        )
                    ]
                else:
                    rows = [r for r in rows if isinstance(r, dict) and row_passes_filters(r, raw_filters)]
        except json.JSONDecodeError:
            pass

    # Sort
    if sort_by:
        reverse = sort_dir == "desc"

        def sort_key(row: dict):
            v = row.get(sort_by)
            try:
                return float(v)
            except (TypeError, ValueError):
                return str(v) if v is not None else ""

        try:
            rows = sorted(rows, key=sort_key, reverse=reverse)
        except Exception:
            pass

    # Build CSV
    sub_fields = [sf for sf in (field.sub_fields or [])]
    headers = [getattr(sf, "key", "") for sf in sub_fields]
    key_to_sf = {sf.key: sf for sf in sub_fields}
    output = BytesIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(
            [
                _serialize_multi_item_cell_for_csv(
                    r.get(key, "") if isinstance(r, dict) else "",
                    getattr(key_to_sf.get(key), "field_type", None),
                )
                for key in headers
            ]
        )

    filename = f"multi_items_{field.key}_{field.id}.csv"
    return StreamingResponse(
        BytesIO(output.getvalue()),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.get("/overview")
async def get_entries_overview(
    year: int = Query(..., ge=2000, le=2100),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List KPIs with entry status and first 2 field preview. For admins, shows assigned user's entry (same source as data entry operator)."""
    org_id = _org_id(current_user, organization_id)
    as_admin = current_user.role.value in ("ORG_ADMIN", "SUPER_ADMIN")
    items = await list_entries_overview(db, current_user.id, org_id, year, as_admin=as_admin)
    return items


@router.get("/available-kpis")
async def get_available_kpis(
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List KPIs the current user can enter data for (assigned KPIs or all org KPIs for admin)."""
    org_id = _org_id(current_user, organization_id)
    kpis = await list_available_kpis(db, current_user.id, org_id)
    return [{"id": k.id, "name": k.name, "year": k.year, "domain_id": k.domain_id} for k in kpis]


@router.get("/kpi-api-info")
async def get_kpi_api_info(
    kpi_id: int = Query(...),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return KPI entry metadata + edit capabilities.

    - can_edit: user can edit at least one field (field-level or KPI-level data_entry)
    - kpi_level_can_edit: user has KPI-level data_entry right (ignores row/field-only grants)
    """
    org_id = _org_id(current_user, organization_id)
    can_view = await user_can_view_kpi(db, current_user.id, kpi_id)
    if not can_view:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not assigned to this KPI")
    field_access = await get_user_field_access_for_kpi(db, current_user.id, kpi_id)
    if field_access is None:
        can_edit = await user_can_edit_kpi(db, current_user.id, kpi_id)
    else:
        can_edit = any(perm == "data_entry" for perm in field_access.values())
    if current_user.role.value in ("ORG_ADMIN", "SUPER_ADMIN"):
        kpi_level_can_edit = True
    else:
        kpi_level_res = await db.execute(
            select(KpiRoleAssignment.id)
            .join(
                UserOrganizationRole,
                UserOrganizationRole.organization_role_id == KpiRoleAssignment.organization_role_id,
            )
            .where(
                UserOrganizationRole.user_id == current_user.id,
                KpiRoleAssignment.kpi_id == kpi_id,
                KpiRoleAssignment.assignment_type == "data_entry",
            )
            .limit(1)
        )
        kpi_level_can_edit = kpi_level_res.scalar_one_or_none() is not None
    res = await db.execute(select(KPI).where(KPI.id == kpi_id, KPI.organization_id == org_id))
    kpi = res.scalar_one_or_none()
    if not kpi:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="KPI not found")
    return {
        "entry_mode": getattr(kpi, "entry_mode", None) or "manual",
        "api_endpoint_url": getattr(kpi, "api_endpoint_url", None),
        "can_edit": can_edit,
        "kpi_level_can_edit": kpi_level_can_edit,
    }


@router.get("/multi-items/add-row-info")
async def get_multi_items_add_row_info(
    field_id: int = Query(..., description="Multi-line items field ID"),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return whether current user can add rows to the given multi-line field."""
    org_id = _org_id(current_user, organization_id)
    field = await _load_multi_items_field(db, org_id, field_id)
    if not field:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")
    can_view = await user_can_view_kpi(db, current_user.id, field.kpi_id, org_id=org_id)
    if not can_view:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to view this KPI")
    can_add = await user_can_add_row_multi_line_field(db, current_user.id, field.kpi_id, field.id)
    return {"can_add_row": bool(can_add)}


@router.get("/multi-items/page-context", response_model=MultiItemsPageContextResponse)
async def get_multi_items_page_context(
    kpi_id: int = Query(..., description="KPI id for the entry"),
    year: int = Query(..., ge=2000, le=2100),
    field_id: int = Query(..., description="Multi-line items field id"),
    period_key: str | None = Query(None, description="Period key: '', H1, H2, Q1-Q4, 01-12"),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Multi-line full page view context in a single call:
    - entry_id (created if missing)
    - KPI minimal name
    - the requested multi-line field (with sub_fields + permissions)
    - can_edit and KPI-level edit
    - can_add_row for this multi-line field
    """
    org_id = _org_id(current_user, organization_id)
    can_view = await user_can_view_kpi(db, current_user.id, kpi_id, org_id=org_id)
    if not can_view:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No access to this KPI")

    pk = (period_key or "").strip()[:8]
    entry, created = await get_or_create_entry(db, current_user.id, org_id, kpi_id, year, period_key=pk)
    if not entry:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="KPI not in organization")
    if created:
        await db.commit()

    # KPI minimal
    kpi_res = await db.execute(select(KPI.id, KPI.name).where(KPI.id == kpi_id, KPI.organization_id == org_id))
    kpi_row = kpi_res.first()
    if not kpi_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="KPI not found")
    kpi_name = str(kpi_row[1] or "")

    # Load just the requested field (fast path) + permissions payload like /entries/fields
    field_res = await db.execute(
        select(KPIField)
        .join(KPI, KPI.id == KPIField.kpi_id)
        .where(KPIField.id == int(field_id), KPIField.kpi_id == int(kpi_id), KPI.organization_id == org_id)
        .options(selectinload(KPIField.sub_fields))
    )
    f = field_res.scalar_one_or_none()
    if not f or getattr(f, "field_type", None) != FieldType.multi_line_items:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Multi-item field not found")

    field_access = await get_user_field_access_for_kpi(db, current_user.id, kpi_id)
    if field_access is None:
        can_edit = await user_can_edit_kpi(db, current_user.id, kpi_id, org_id=org_id)
        whole_perm = "data_entry" if can_edit else "view"
    else:
        can_edit = any(perm == "data_entry" for perm in field_access.values())
        whole_perm = field_access.get((f.id, None))

    # KPI-level can_edit (same semantics as /entries/kpi-api-info)
    if current_user.role.value in ("ORG_ADMIN", "SUPER_ADMIN"):
        kpi_level_can_edit = True
    else:
        kpi_level_res = await db.execute(
            select(KpiRoleAssignment.id)
            .join(
                UserOrganizationRole,
                UserOrganizationRole.organization_role_id == KpiRoleAssignment.organization_role_id,
            )
            .where(
                UserOrganizationRole.user_id == current_user.id,
                KpiRoleAssignment.kpi_id == kpi_id,
                KpiRoleAssignment.assignment_type == "data_entry",
            )
            .limit(1)
        )
        kpi_level_can_edit = kpi_level_res.scalar_one_or_none() is not None

    # Field + sub-field permissions
    sub_fields = list(getattr(f, "sub_fields", None) or [])
    sub_payload: list[MultiItemsPageContextSubField] = []
    if field_access is None:
        for s in sub_fields:
            sub_payload.append(
                MultiItemsPageContextSubField(
                    id=getattr(s, "id", None),
                    key=str(getattr(s, "key", "")),
                    name=str(getattr(s, "name", getattr(s, "key", ""))),
                    field_type=getattr(getattr(s, "field_type", None), "value", getattr(s, "field_type", None)),
                    is_required=bool(getattr(s, "is_required", False)),
                    sort_order=getattr(s, "sort_order", None),
                    config=getattr(s, "config", None),
                    can_view=True,
                    can_edit=bool(can_edit),
                )
            )
    else:
        for s in sub_fields:
            sid = getattr(s, "id", None)
            if sid is None:
                continue
            sub_perm = field_access.get((f.id, int(sid))) or whole_perm
            if not sub_perm:
                continue
            sub_payload.append(
                MultiItemsPageContextSubField(
                    id=int(sid),
                    key=str(getattr(s, "key", "")),
                    name=str(getattr(s, "name", getattr(s, "key", ""))),
                    field_type=getattr(getattr(s, "field_type", None), "value", getattr(s, "field_type", None)),
                    is_required=bool(getattr(s, "is_required", False)),
                    sort_order=getattr(s, "sort_order", None),
                    config=getattr(s, "config", None),
                    can_view=sub_perm in ("view", "data_entry"),
                    can_edit=sub_perm == "data_entry",
                )
            )

    out_field = MultiItemsPageContextField(
        id=int(f.id),
        kpi_id=int(f.kpi_id),
        name=str(f.name or ""),
        key=str(f.key or ""),
        field_type=getattr(getattr(f, "field_type", None), "value", str(getattr(f, "field_type", ""))),
        full_page_multi_items=bool(getattr(f, "full_page_multi_items", False)),
        row_level_user_access_enabled=bool(getattr(f, "row_level_user_access_enabled", False)),
        config=getattr(f, "config", None),
        sub_fields=sub_payload,
    )

    can_add = await user_can_add_row_multi_line_field(db, current_user.id, kpi_id, int(f.id))

    return MultiItemsPageContextResponse(
        entry_id=int(entry.id),
        kpi_id=int(kpi_id),
        kpi_name=kpi_name,
        field=out_field,
        can_edit=bool(can_edit),
        kpi_level_can_edit=bool(kpi_level_can_edit),
        can_add_row=bool(can_add),
    )


@router.post("/sync-from-api")
async def entry_sync_from_api(
    kpi_id: int = Query(...),
    year: int = Query(..., ge=2000, le=2100),
    organization_id: int | None = Query(None),
    sync_mode: str = Query(
        "override",
        description="override = replace multi-line rows; append = append; upsert = merge by upsert_match_keys",
        pattern="^(override|append|upsert)$",
    ),
    upsert_match_keys: str | None = Query(
        None,
        description="JSON: multi_line field key -> sub_field key (each multi-line table when sync_mode=upsert)",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Fetch entry data from the KPI's API endpoint. User must be allowed to edit this KPI. UI sync_mode wins; API override_existing is ignored."""
    org_id = _org_id(current_user, organization_id)
    can = await user_can_edit_kpi(db, current_user.id, kpi_id)
    if not can:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to edit this KPI")
    parsed = parse_upsert_match_keys_json(upsert_match_keys)
    result = await sync_kpi_entry_from_api(
        db,
        kpi_id,
        org_id,
        year,
        current_user.id,
        sync_mode=sync_mode,
        upsert_match_keys=parsed,
    )
    if result is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="KPI not found, not in API mode, API endpoint not set, or API call failed",
        )
    await db.commit()
    return result


@router.get("/fields")
async def get_entry_fields(
    kpi_id: int = Query(...),
    field_id: int | None = Query(None, description="Optional: return only this field id (fast path for entry pages)"),
    minimal: bool = Query(False, description="When true, omit options and other non-essential payload."),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List fields for a KPI the current user can view or enter data for.
    Only returns fields the user has at least view access to. Each field (and sub_field) includes can_view and can_edit."""
    org_id = _org_id(current_user, organization_id)
    can_view_kpi = await user_can_view_kpi(db, current_user.id, kpi_id, org_id=org_id)
    if not can_view_kpi:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not assigned to this KPI")
    if field_id is not None:
        # True fast path: load only the requested field (and its options/sub_fields) from DB.
        # Avoid loading all fields for the KPI and filtering in Python.
        load_options = not minimal
        field_res = await db.execute(
            select(KPIField)
            .join(KPI, KPI.id == KPIField.kpi_id)
            .where(
                KPIField.id == int(field_id),
                KPIField.kpi_id == int(kpi_id),
                KPI.organization_id == org_id,
            )
            .options(
                *( [selectinload(KPIField.options)] if load_options else [] ),
                selectinload(KPIField.sub_fields),
            )
        )
        one = field_res.scalar_one_or_none()
        fields = [one] if one is not None else []
    else:
        fields = await list_kpi_fields_service(db, kpi_id, org_id)
    field_access = await get_user_field_access_for_kpi(db, current_user.id, kpi_id)

    if field_access is None:
        can_edit_kpi = await user_can_edit_kpi(db, current_user.id, kpi_id, org_id=org_id)
        return [
            {
                "id": f.id,
                "kpi_id": f.kpi_id,
                "name": f.name,
                "key": f.key,
                "field_type": f.field_type.value,
                "formula_expression": f.formula_expression,
                "is_required": f.is_required,
                "sort_order": f.sort_order,
                "config": getattr(f, "config", None),
                "carry_forward_data": getattr(f, "carry_forward_data", False),
                "full_page_multi_items": getattr(f, "full_page_multi_items", False),
                "row_level_user_access_enabled": getattr(f, "row_level_user_access_enabled", False),
                "can_view": True,
                "can_edit": can_edit_kpi,
                "options": [] if minimal else [{"value": o.value, "label": o.label} for o in (f.options or [])],
                "sub_fields": [
                    {"id": s.id, "field_id": s.field_id, "name": s.name, "key": s.key, "field_type": s.field_type.value, "is_required": s.is_required, "sort_order": s.sort_order, "config": getattr(s, "config", None), "can_view": True, "can_edit": can_edit_kpi}
                    for s in (getattr(f, "sub_fields", None) or [])
                ],
            }
            for f in fields
        ]

    out = []
    for f in fields:
        sub_fields_list = getattr(f, "sub_fields", None) or []
        whole_perm = field_access.get((f.id, None))
        if whole_perm:
            can_view_f = whole_perm in ("view", "data_entry")
            can_edit_f = whole_perm == "data_entry"
        else:
            can_view_f = any(field_access.get((f.id, s.id)) in ("view", "data_entry") for s in sub_fields_list)
            can_edit_f = any(field_access.get((f.id, s.id)) == "data_entry" for s in sub_fields_list)
        if not can_view_f and not can_edit_f:
            continue
        sub_payload = []
        for s in sub_fields_list:
            sub_perm = field_access.get((f.id, s.id)) or whole_perm
            if not sub_perm:
                continue
            sub_payload.append({
                "id": s.id, "field_id": s.field_id, "name": s.name, "key": s.key,
                "field_type": s.field_type.value, "is_required": s.is_required, "sort_order": s.sort_order,
                "config": getattr(s, "config", None),
                "can_view": sub_perm in ("view", "data_entry"),
                "can_edit": sub_perm == "data_entry",
            })
        if f.field_type == FieldType.multi_line_items and not whole_perm and not sub_payload:
            continue
        if f.field_type != FieldType.multi_line_items and not whole_perm:
            continue
        out.append({
            "id": f.id,
            "kpi_id": f.kpi_id,
            "name": f.name,
            "key": f.key,
            "field_type": f.field_type.value,
            "formula_expression": f.formula_expression,
            "is_required": f.is_required,
            "sort_order": f.sort_order,
            "config": getattr(f, "config", None),
            "carry_forward_data": getattr(f, "carry_forward_data", False),
            "full_page_multi_items": getattr(f, "full_page_multi_items", False),
            "row_level_user_access_enabled": getattr(f, "row_level_user_access_enabled", False),
            "can_view": whole_perm in ("view", "data_entry") if whole_perm else bool(sub_payload),
            "can_edit": whole_perm == "data_entry" if whole_perm else any(sp.get("can_edit") for sp in sub_payload),
            "options": [] if minimal else [{"value": o.value, "label": o.label} for o in (f.options or [])],
            "sub_fields": sub_payload if sub_payload else [
                {"id": s.id, "field_id": s.field_id, "name": s.name, "key": s.key, "field_type": s.field_type.value, "is_required": s.is_required, "sort_order": s.sort_order, "config": getattr(s, "config", None), "can_view": True, "can_edit": whole_perm == "data_entry"}
                for s in sub_fields_list
            ] if whole_perm else [],
        })
    return out


def _excel_sheet_name(name: str, max_len: int = 31) -> str:
    """Sanitize sheet name for Excel (max 31 chars; no : \\ / ? * [ ])."""
    invalid = ":\\/?*[]"
    out = "".join(c if c not in invalid else "_" for c in (name or "Sheet"))
    return out[:max_len].strip() or "Sheet"


async def _build_kpi_entry_xlsx(
    kpi_name: str,
    year: int,
    org_id: int,
    fields: list,
    entry,
) -> bytes:
    """Build Excel workbook: one sheet for scalar fields, one sheet per multi_line_items field."""
    from openpyxl import Workbook

    wb = Workbook()
    value_by_field_id = {}
    if entry and getattr(entry, "field_values", None):
        for fv in entry.field_values:
            value_by_field_id[fv.field_id] = fv

    # --- Sheet 1: Scalar (and formula) fields ---
    scalar_types = (
        FieldType.single_line_text,
        FieldType.multi_line_text,
        FieldType.number,
        FieldType.date,
        FieldType.boolean,
        FieldType.attachment,
        FieldType.formula,
        FieldType.reference,
        FieldType.multi_reference,
        FieldType.mixed_list,
    )
    scalar_fields = [f for f in fields if getattr(f, "field_type", None) in scalar_types]
    ws_scalar = wb.active
    ws_scalar.title = _excel_sheet_name("KPI Data")
    ws_scalar.append(["Field", "Value"])
    for f in scalar_fields:
        fv = value_by_field_id.get(f.id)
        if fv is None:
            ws_scalar.append([f.name, ""])
            continue
        ft = getattr(f, "field_type", None)
        if ft == FieldType.multi_reference and isinstance(fv.value_json, list):
            val = "; ".join(str(x) for x in fv.value_json)
        elif ft == FieldType.mixed_list and isinstance(fv.value_json, list):
            val = "; ".join(str(x) for x in fv.value_json)
        elif fv.value_text is not None:
            val = fv.value_text
        elif fv.value_number is not None:
            val = fv.value_number
        elif fv.value_boolean is not None:
            val = "Yes" if fv.value_boolean else "No"
        elif fv.value_date is not None:
            val = str(fv.value_date)[:10] if fv.value_date else ""
        elif fv.value_json is not None:
            val = str(fv.value_json)[:500]
        else:
            val = ""
        ws_scalar.append([f.name, val])

    # --- One sheet per multi_line_items field ---
    multi_fields = [f for f in fields if getattr(f, "field_type", None) == FieldType.multi_line_items]
    for idx, f in enumerate(multi_fields):
        sub_fields = list(getattr(f, "sub_fields", None) or [])
        keys = [s.key for s in sub_fields]
        if not keys:
            keys = ["value"]
        sheet_name = _excel_sheet_name(f.name) or f"Table_{idx + 1}"
        if sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = _excel_sheet_name(f"{f.name}_{idx}") or f"Table_{idx + 1}"
        ws = wb.create_sheet(title=sheet_name)
        ws.append(keys)
        pairs = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=f)
        rows = [r for _, r in pairs] if pairs else []
        key_to_sf = {s.key: s for s in sub_fields}
        for row in rows:
            if not isinstance(row, dict):
                ws.append([""] * len(keys))
                continue

            def _cell_out(col_key: str):
                raw = row.get(col_key, "")
                sf = key_to_sf.get(col_key)
                if sf and getattr(sf, "field_type", None) == FieldType.multi_reference and isinstance(raw, list):
                    return "; ".join(str(x) for x in raw)
                if sf and getattr(sf, "field_type", None) == FieldType.mixed_list and isinstance(raw, list):
                    return "; ".join(str(x) for x in raw)
                return raw if raw is not None else ""

            ws.append([_cell_out(k) for k in keys])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export-excel")
async def export_entry_excel(
    kpi_id: int = Query(...),
    year: int = Query(..., ge=2000, le=2100),
    period_key: str = Query("", description="Period: '', H1, H2, Q1-Q4, 01-12 for time dimension"),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download KPI entry data as Excel: scalar fields in one sheet, each multi_line_items in its own sheet."""
    org_id = _org_id(current_user, organization_id)
    can_view = await user_can_view_kpi(db, current_user.id, kpi_id)
    if not can_view:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No access to this KPI")
    kpi_res = await db.execute(
        select(KPI).where(KPI.id == kpi_id, KPI.organization_id == org_id)
    )
    kpi = kpi_res.scalar_one_or_none()
    if not kpi:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="KPI not found")
    fields = await list_kpi_fields_service(db, kpi_id, org_id)
    pk = (period_key or "").strip()[:8]
    entry_res = await db.execute(
        select(KPIEntry)
        .where(
            KPIEntry.kpi_id == kpi_id,
            KPIEntry.organization_id == org_id,
            KPIEntry.year == year,
            KPIEntry.period_key == pk,
        )
        .options(selectinload(KPIEntry.field_values).selectinload(KPIFieldValue.field))
    )
    entry = entry_res.scalar_one_or_none()
    xlsx_bytes = await _build_kpi_entry_xlsx(
        kpi_name=getattr(kpi, "name", "") or f"KPI_{kpi_id}",
        year=year,
        org_id=org_id,
        fields=fields,
        entry=entry,
    )
    filename = f"KPI_{kpi_id}_{year}_org{org_id}.xlsx"
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_kpi_entry_xlsx(
    content: bytes,
    fields: list,
) -> dict[int, dict]:
    """Parse uploaded Excel into field values. Returns {field_id: {value_text, value_number, value_boolean, value_date, value_json}}."""
    from openpyxl import load_workbook

    from app.entries.service import coerce_multi_reference_raw

    wb = load_workbook(filename=BytesIO(content), data_only=True)
    result: dict[int, dict] = {}

    # Build lookup from field name -> field
    name_to_field = {f.name.strip().lower(): f for f in fields}
    sheet_name_to_field = {}
    for f in fields:
        if getattr(f, "field_type", None) == FieldType.multi_line_items:
            sanitized = _excel_sheet_name(f.name).lower()
            sheet_name_to_field[sanitized] = f

    # Parse scalar sheet ("KPI Data")
    scalar_sheet = None
    for ws in wb.worksheets:
        if ws.title.lower() == "kpi data":
            scalar_sheet = ws
            break
    if not scalar_sheet and wb.worksheets:
        scalar_sheet = wb.worksheets[0]

    if scalar_sheet:
        rows = list(scalar_sheet.iter_rows(values_only=True))
        for row in rows[1:]:  # skip header
            if not row or len(row) < 2:
                continue
            field_name = str(row[0]).strip().lower() if row[0] else ""
            raw_value = row[1]
            field = name_to_field.get(field_name)
            if not field:
                continue
            ft = getattr(field, "field_type", None)
            if ft == FieldType.formula:
                continue  # skip formula fields on upload
            if ft == FieldType.multi_line_items:
                continue  # handled separately
            val: dict = {}
            if raw_value is None or raw_value == "":
                pass
            elif ft == FieldType.number:
                try:
                    val["value_number"] = float(raw_value)
                except (TypeError, ValueError):
                    val["value_text"] = str(raw_value)
            elif ft == FieldType.boolean:
                if isinstance(raw_value, bool):
                    val["value_boolean"] = raw_value
                else:
                    s = str(raw_value).strip().lower()
                    val["value_boolean"] = s in ("1", "true", "yes", "y")
            elif ft == FieldType.date:
                if hasattr(raw_value, "date"):
                    try:
                        val["value_date"] = raw_value.date().isoformat()
                    except Exception:
                        val["value_text"] = str(raw_value)
                else:
                    val["value_text"] = str(raw_value)
            elif ft == FieldType.multi_reference:
                parsed = coerce_multi_reference_raw(str(raw_value) if raw_value is not None else "")
                val["value_json"] = parsed if parsed else None
            elif ft == FieldType.mixed_list:
                parsed = coerce_mixed_list_raw(str(raw_value) if raw_value is not None else "")
                val["value_json"] = parsed if parsed else None
            else:
                val["value_text"] = str(raw_value) if raw_value is not None else None
            result[field.id] = val

    # Parse multi_line_items sheets
    for ws in wb.worksheets:
        title_lower = ws.title.lower()
        if title_lower == "kpi data":
            continue
        field = sheet_name_to_field.get(title_lower)
        if not field:
            # Try matching by prefix
            for sn, f in sheet_name_to_field.items():
                if title_lower.startswith(sn[:20]):
                    field = f
                    break
        if not field:
            continue
        sub_fields = list(getattr(field, "sub_fields", None) or [])
        key_to_sf = {s.key: s for s in sub_fields}
        name_to_sf = {s.name.strip().lower(): s for s in sub_fields}
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c else "" for c in rows[0]]
        col_to_key: list[str | None] = []
        for col in header:
            col_lower = col.lower()
            if col in key_to_sf:
                col_to_key.append(col)
            elif col_lower in name_to_sf:
                col_to_key.append(name_to_sf[col_lower].key)
            else:
                col_to_key.append(None)
        items: list[dict] = []
        for row in rows[1:]:
            if not row:
                continue
            item: dict = {}
            empty = True
            for i, raw in enumerate(row):
                if i >= len(col_to_key):
                    continue
                key = col_to_key[i]
                if not key:
                    continue
                if raw is None or raw == "":
                    continue
                empty = False
                sf = key_to_sf.get(key)
                sf_type = sf.field_type if sf else "single_line_text"
                if sf_type == FieldType.number or sf_type == "number":
                    try:
                        item[key] = float(raw)
                    except (TypeError, ValueError):
                        item[key] = str(raw)
                elif sf_type == FieldType.boolean or sf_type == "boolean":
                    if isinstance(raw, bool):
                        item[key] = raw
                    else:
                        s = str(raw).strip().lower()
                        item[key] = s in ("1", "true", "yes", "y")
                elif sf_type == FieldType.date or sf_type == "date":
                    if hasattr(raw, "date"):
                        try:
                            item[key] = raw.date().isoformat()
                        except Exception:
                            item[key] = str(raw)
                    else:
                        item[key] = str(raw)
                elif sf_type == FieldType.multi_reference or sf_type == "multi_reference":
                    parsed = coerce_multi_reference_raw(str(raw) if raw is not None else "")
                    item[key] = parsed if parsed else None
                elif sf_type == FieldType.mixed_list or sf_type == "mixed_list":
                    parsed = coerce_mixed_list_raw(str(raw) if raw is not None else "")
                    item[key] = parsed if parsed else None
                else:
                    item[key] = str(raw)
            if not empty:
                items.append(item)
        result[field.id] = {"value_json": items}

    return result


@router.post("/import-excel")
async def import_entry_excel(
    kpi_id: int = Query(...),
    year: int = Query(..., ge=2000, le=2100),
    period_key: str = Query("", description="Period: '', H1, H2, Q1-Q4, 01-12 for time dimension"),
    organization_id: int | None = Query(None),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload KPI entry data from Excel (same format as download). Auth: Org Admin or data_entry for this KPI."""
    org_id = _org_id(current_user, organization_id)
    can_edit = await user_can_edit_kpi(db, current_user.id, kpi_id)
    if not can_edit:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No edit access to this KPI")
    kpi_res = await db.execute(
        select(KPI).where(KPI.id == kpi_id, KPI.organization_id == org_id)
    )
    kpi = kpi_res.scalar_one_or_none()
    if not kpi:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="KPI not found")
    fields = await list_kpi_fields_service(db, kpi_id, org_id)
    content = await file.read()
    try:
        parsed = _parse_kpi_entry_xlsx(content, fields)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to parse Excel: {e}")
    # Get or create entry for this period
    entry, _ = await get_or_create_entry(db, current_user.id, org_id, kpi_id, year, period_key=(period_key or "").strip()[:8])
    if not entry:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="KPI not in organization")
    if entry.is_locked:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Entry is locked")
    # Build FieldValueInput list
    from app.entries.schemas import FieldValueInput
    values = []
    for field in fields:
        if field.id not in parsed:
            continue
        data = parsed[field.id]
        values.append(
            FieldValueInput(
                field_id=field.id,
                value_text=data.get("value_text"),
                value_number=data.get("value_number"),
                value_boolean=data.get("value_boolean"),
                value_date=data.get("value_date"),
                value_json=data.get("value_json"),
            )
        )
    try:
        await save_entry_values(db, entry.id, current_user.id, values, kpi_id, org_id)
    except EntryValidationError:
        raise  # Handled by app exception_handler; returns 400 with errors list
    await db.commit()
    await db.refresh(entry)
    return {"message": "Import successful", "entry_id": entry.id, "fields_updated": len(values)}


def _reverse_ref_tokens_from_cell(cell) -> list[tuple[str, str]]:
    """Return (display_label, normalized_token) pairs for reverse-reference matching."""
    if cell is None or cell == "":
        return []
    out: list[tuple[str, str]] = []
    if isinstance(cell, list):
        for x in cell:
            if x is None:
                continue
            label = str(x).strip()
            if not label:
                continue
            t = _normalize_reference_value(label)
            if t:
                out.append((label, t))
        return out
    label = str(cell).strip()
    t = _normalize_reference_value(label)
    if t:
        out.append((label, t))
    return out


@router.get("/reverse-references")
async def get_reverse_references_for_entry(
    kpi_id: int = Query(..., description="Parent KPI id"),
    entry_id: int = Query(..., description="Parent KPI entry id"),
    include_rows: bool = Query(
        False,
        description="When false, return only KPI names + value tokens/counts without embedding row data.",
    ),
    values_limit: int = Query(
        500,
        ge=1,
        le=5000,
        description="When include_rows is false, maximum number of dropdown values to return per related KPI.",
    ),
    values_search: str | None = Query(
        None,
        description="When include_rows is false, optional case-insensitive search over value token/label.",
    ),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    For a given parent KPI entry, return reverse-reference info for reference sub-fields in multi_line_items fields.
    Response shape (list of tabs):
    [
      {
        "child_kpi_id": int,
        "child_kpi_name": str,
        "values": [ { "token": str, "label": str, "count": int } ],
        "rows": [
          {
            "entry_id": int,
            "year": int,
            "period_key": str,
            "value_token": str,
            "value_display": str,
            "child_field_id": int,
            "child_field_key": str,
            "child_field_name": str,
            "child_sub_field_key": str,
            "child_sub_field_name": str,
            "row_index": int,
            "row": dict,
          },
          ...
        ],
        "sub_fields": [ { "key": str, "name": str } ]
      },
      ...
    ]
    """
    org_id = _org_id(current_user, organization_id)
    can_view = await user_can_view_kpi(db, current_user.id, kpi_id)
    if not can_view:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No access to this KPI")

    # Load parent entry with its field values and fields
    parent_entry_res = await db.execute(
        select(KPIEntry)
        .where(
            KPIEntry.id == entry_id,
            KPIEntry.organization_id == org_id,
            KPIEntry.kpi_id == kpi_id,
        )
        .options(
            selectinload(KPIEntry.field_values).selectinload(KPIFieldValue.field).selectinload(KPIField.sub_fields),
        )
    )
    parent_entry = parent_entry_res.scalar_one_or_none()
    if not parent_entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entry not found")

    parent_year = parent_entry.year
    parent_period_key = (getattr(parent_entry, "period_key", None) or "").strip()[:8]

    # Load parent KPI with fields for lookup by key
    parent_kpi_res = await db.execute(
        select(KPI).where(KPI.id == kpi_id).options(selectinload(KPI.fields))
    )
    parent_kpi = parent_kpi_res.scalar_one_or_none()
    if not parent_kpi:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="KPI not found")

    # Resolve effective time dimension: KPI-level if set, else org-level (for filtering and display)
    org_row = await db.get(Organization, org_id)
    org_td_raw = getattr(org_row, "time_dimension", None) or "yearly"
    try:
        org_td = TimeDimension(org_td_raw)
    except ValueError:
        org_td = TimeDimension.YEARLY
    parent_kpi_td_raw = getattr(parent_kpi, "time_dimension", None)
    parent_kpi_td = None
    if parent_kpi_td_raw:
        try:
            parent_kpi_td = TimeDimension(parent_kpi_td_raw)
        except ValueError:
            pass
    effective_td = effective_kpi_time_dimension(parent_kpi_td, org_td)
    parent_fields_by_key = {f.key: f for f in getattr(parent_kpi, "fields", []) or []}
    parent_fv_by_field_id = {fv.field_id: fv for fv in getattr(parent_entry, "field_values", []) or []}

    # Build descriptors for reference sub-fields that point to this parent KPI.
    # Critical for performance: do NOT load all fields+sub_fields in the org.
    sub_fields_res = await db.execute(
        select(
            KPIField.kpi_id,
            KPIField.id,
            KPIField.key,
            KPIField.name,
            KPIFieldSubField.id,
            KPIFieldSubField.key,
            KPIFieldSubField.name,
            KPIFieldSubField.config,
        )
        .select_from(KPIFieldSubField)
        .join(KPIField, KPIField.id == KPIFieldSubField.field_id)
        .join(KPI, KPI.id == KPIField.kpi_id)
        .where(
            KPI.organization_id == org_id,
            KPIFieldSubField.field_type.in_([FieldType.reference, FieldType.multi_reference]),
        )
    )

    descriptors_by_child_kpi: dict[int, list[dict]] = {}
    for child_kpi_id, child_field_id, child_field_key, child_field_name, child_sub_field_id, child_sub_field_key, child_sub_field_name, cfg in sub_fields_res.all():
        cfg = cfg or {}
        sid = cfg.get("reference_source_kpi_id")
        if not sid or int(sid) != kpi_id:
            continue
        skey = cfg.get("reference_source_field_key")
        if not skey:
            continue
        parent_sub_key = cfg.get("reference_source_sub_field_key")
        desc = {
            "child_kpi_id": int(child_kpi_id),
            "child_field_id": int(child_field_id),
            "child_field_key": str(child_field_key),
            "child_field_name": str(child_field_name),
            "child_sub_field_id": int(child_sub_field_id),
            "child_sub_field_key": str(child_sub_field_key),
            "child_sub_field_name": str(child_sub_field_name),
            "parent_field_key": str(skey),
            "parent_sub_field_key": str(parent_sub_key) if parent_sub_key else None,
        }
        descriptors_by_child_kpi.setdefault(int(child_kpi_id), []).append(desc)

    if not descriptors_by_child_kpi:
        return {
            "time_filter": {
                "year": parent_year,
                "period_key": parent_period_key,
                "effective_time_dimension": effective_td.value if hasattr(effective_td, "value") else str(effective_td),
            },
            "tabs": [],
        }

    # Load child KPI names
    child_kpi_ids = list(descriptors_by_child_kpi.keys())
    kpi_res = await db.execute(select(KPI).where(KPI.id.in_(child_kpi_ids)))
    child_kpis = {k.id: k for k in kpi_res.scalars().all()}

    response_tabs: list[dict] = []

    # Fast path for the UI: include_rows=false only needs token counts and deep-link defaults.
    # Avoid scanning all child multi_line rows in Python and avoid returning 20k dropdown options in one response.
    if not include_rows:
        parent_multi_cache: dict[int, list[tuple[int, dict]]] = {}

        for child_kpi_id, desc_list in descriptors_by_child_kpi.items():
            child_kpi = child_kpis.get(child_kpi_id)
            child_kpi_name = getattr(child_kpi, "name", f"KPI #{child_kpi_id}")

            # Compute parent tokens (token -> label) once for this child KPI's descriptors.
            descriptor_tokens: dict[str, dict[str, str]] = {}
            for d in desc_list:
                p_key = d["parent_field_key"]
                p_sub = d["parent_sub_field_key"]
                parent_field = parent_fields_by_key.get(p_key)
                if not parent_field:
                    continue
                fv = parent_fv_by_field_id.get(parent_field.id)
                if not fv:
                    continue
                if p_sub:
                    # Parent is a multi_line_items field: load from relational rows (cached per field id).
                    if parent_field.id not in parent_multi_cache:
                        parent_multi_cache[parent_field.id] = await _load_multi_line_row_dicts(
                            db, entry_id=parent_entry.id, field=parent_field
                        )
                    items = [r for _, r in parent_multi_cache[parent_field.id]] if parent_multi_cache[parent_field.id] else []
                    for row in items:
                        if not isinstance(row, dict):
                            continue
                        cell = row.get(p_sub)
                        for label, token in _reverse_ref_tokens_from_cell(cell):
                            descriptor_tokens.setdefault(token, {"label": label})
                else:
                    p_ft = getattr(parent_field, "field_type", None)
                    if p_ft == FieldType.multi_reference:
                        arr = fv.value_json if isinstance(fv.value_json, list) else []
                        for x in arr:
                            label = str(x).strip() if x is not None else ""
                            token = _normalize_reference_value(label)
                            if token:
                                descriptor_tokens.setdefault(token, {"label": label})
                    else:
                        raw_val = None
                        if fv.value_text not in (None, ""):
                            raw_val = fv.value_text
                        elif fv.value_number is not None:
                            raw_val = fv.value_number
                        elif fv.value_boolean is not None:
                            raw_val = fv.value_boolean
                        elif fv.value_date is not None:
                            raw_val = fv.value_date.isoformat()
                        if raw_val is not None:
                            label = (str(raw_val)).strip()
                            token = _normalize_reference_value(label)
                            if token:
                                descriptor_tokens.setdefault(token, {"label": label})

            if not descriptor_tokens:
                continue

            # Count matches in SQL per descriptor and merge results.
            # IMPORTANT: Do NOT pass 20k tokens in an IN (...) list (very slow / huge query).
            # Instead, aggregate counts for child tokens then intersect with parent tokens in Python.
            token_counts: dict[str, int] = {}
            for d in desc_list:
                child_field_id = int(d["child_field_id"])
                child_sub_field_id = int(d["child_sub_field_id"])

                # Find the single child entry id for this KPI+year+period in this org (1 row).
                child_entry_id_res = await db.execute(
                    select(KPIEntry.id).where(
                        KPIEntry.organization_id == org_id,
                        KPIEntry.kpi_id == child_kpi_id,
                        KPIEntry.year == parent_year,
                        KPIEntry.period_key == parent_period_key,
                    )
                )
                child_entry_id = child_entry_id_res.scalar_one_or_none()
                if child_entry_id is None:
                    continue

                # Normalize reference token in SQL similar to _normalize_reference_value (strip + lower),
                # but also keep the original-case label for UI + deep links.
                cell = KpiMultiLineCell
                raw_expr = func.trim(
                    func.coalesce(
                        cast(cell.value_text, String()),
                        func.json_extract_path_text(cell.value_json, "label"),
                        func.json_extract_path_text(cell.value_json, "value"),
                        func.json_extract_path_text(cell.value_json, "token"),
                        cast(cell.value_json, String()),
                    )
                )
                token_expr = func.lower(raw_expr)
                label_expr = raw_expr

                base_q = (
                    select(
                        token_expr.label("t"),
                        func.min(label_expr).label("label"),
                        func.count().label("c"),
                    )
                    .select_from(KpiMultiLineRow)
                    .join(KpiMultiLineCell, KpiMultiLineCell.row_id == KpiMultiLineRow.id)
                    .where(
                        KpiMultiLineRow.entry_id == int(child_entry_id),
                        KpiMultiLineRow.field_id == child_field_id,
                        KpiMultiLineCell.sub_field_id == child_sub_field_id,
                        token_expr.isnot(None),
                        token_expr != "",
                    )
                    .group_by(token_expr)
                )

                # Optional server-side search to keep payload small.
                if values_search and values_search.strip():
                    q = values_search.strip()
                    base_q = base_q.where(token_expr.ilike(f"%{q.lower()}%"))

                q = (
                    base_q.order_by(token_expr.asc()).limit(int(values_limit))
                )
                res = await db.execute(q)
                for t, label, c in res.all():
                    if not t:
                        continue
                    token_counts[str(t)] = token_counts.get(str(t), 0) + int(c or 0)
                    # Prefer a real original-case label from child rows.
                    if label:
                        descriptor_tokens.setdefault(str(t), {"label": str(label)})

            if not token_counts:
                continue

            # Build payload from tokens that actually exist in the child KPI (keeps response small).
            values_payload = []
            for token, count in token_counts.items():
                meta = descriptor_tokens.get(token)
                label = meta["label"] if meta else token
                if int(count or 0) <= 0:
                    continue
                values_payload.append({"token": token, "label": label, "count": int(count)})
            if not values_payload:
                continue
            values_payload.sort(key=lambda x: x["label"])

            default_target = desc_list[0] if desc_list else None
            response_tabs.append(
                {
                    "child_kpi_id": child_kpi_id,
                    "child_kpi_name": child_kpi_name,
                    "child_field_id": default_target.get("child_field_id") if default_target else None,
                    "child_sub_field_key": default_target.get("child_sub_field_key") if default_target else None,
                    "values": values_payload,
                    "rows": [],
                    "sub_fields": [],
                }
            )

        return {
            "time_filter": {
                "year": parent_year,
                "period_key": parent_period_key,
                "effective_time_dimension": effective_td.value if hasattr(effective_td, "value") else str(effective_td),
            },
            "tabs": response_tabs,
        }

    for child_kpi_id, desc_list in descriptors_by_child_kpi.items():
        child_kpi = child_kpis.get(child_kpi_id)
        child_kpi_name = getattr(child_kpi, "name", f"KPI #{child_kpi_id}")

        # For this child KPI, compute all parent tokens that are relevant for each descriptor
        descriptor_tokens: dict[str, dict[str, str]] = {}  # token -> {"label": str}
        for d in desc_list:
            p_key = d["parent_field_key"]
            p_sub = d["parent_sub_field_key"]
            parent_field = parent_fields_by_key.get(p_key)
            if not parent_field:
                continue
            fv = parent_fv_by_field_id.get(parent_field.id)
            if not fv:
                continue
            if p_sub:
                # Parent is a multi_line_items field: load from relational rows.
                parent_field_rows = await _load_multi_line_row_dicts(db, entry_id=parent_entry.id, field=parent_field)
                items = [r for _, r in parent_field_rows] if parent_field_rows else []
                for row in items:
                    if not isinstance(row, dict):
                        continue
                    cell = row.get(p_sub)
                    for label, token in _reverse_ref_tokens_from_cell(cell):
                        descriptor_tokens.setdefault(token, {"label": label})
            else:
                p_ft = getattr(parent_field, "field_type", None)
                if p_ft == FieldType.multi_reference:
                    arr = fv.value_json if isinstance(fv.value_json, list) else []
                    for x in arr:
                        label = str(x).strip() if x is not None else ""
                        token = _normalize_reference_value(label)
                        if token:
                            descriptor_tokens.setdefault(token, {"label": label})
                else:
                    # Scalar parent field: use its primary value_text/number/boolean/date
                    raw_val = None
                    if fv.value_text not in (None, ""):
                        raw_val = fv.value_text
                    elif fv.value_number is not None:
                        raw_val = fv.value_number
                    elif fv.value_boolean is not None:
                        raw_val = fv.value_boolean
                    elif fv.value_date is not None:
                        raw_val = fv.value_date.isoformat()
                    if raw_val is not None:
                        label = (str(raw_val)).strip()
                        token = _normalize_reference_value(label)
                        if token:
                            descriptor_tokens.setdefault(token, {"label": label})

        if not descriptor_tokens:
            continue

        # Load entries for this child KPI in this org, filtered by parent's time dimension (same year and period)
        child_entries_res = await db.execute(
            select(KPIEntry)
            .where(
                KPIEntry.organization_id == org_id,
                KPIEntry.kpi_id == child_kpi_id,
                KPIEntry.year == parent_year,
                KPIEntry.period_key == parent_period_key,
            )
            .options(selectinload(KPIEntry.field_values))
        )
        child_entries = list(child_entries_res.scalars().all())

        # For table headers we can reuse the sub_fields of the first descriptor's field (slow path only).
        # We no longer have `all_fields` loaded; fetch only needed fields.
        child_field_cache: dict[int, KPIField] = {}
        first_desc = desc_list[0]
        first_child_field_id = int(first_desc["child_field_id"])
        if first_child_field_id not in child_field_cache:
            cf_res = await db.execute(
                select(KPIField)
                .join(KPI, KPI.id == KPIField.kpi_id)
                .where(KPIField.id == first_child_field_id, KPI.organization_id == org_id)
                .options(selectinload(KPIField.sub_fields))
            )
            cf = cf_res.scalar_one_or_none()
            if cf is not None:
                child_field_cache[first_child_field_id] = cf
        child_field = child_field_cache.get(first_child_field_id)
        sub_fields_payload = []
        if child_field is not None:
            for sf in getattr(child_field, "sub_fields", []) or []:
                sub_fields_payload.append({"key": sf.key, "name": getattr(sf, "name", sf.key)})

        rows_payload: list[dict] = [] if include_rows else []
        token_counts: dict[str, int] = {t: 0 for t in descriptor_tokens.keys()}
        tokens_set = set(descriptor_tokens.keys())

        # Scan child entries for matching reference sub-field values
        multi_cache: dict[tuple[int, int], list[tuple[int, dict]]] = {}
        for entry in child_entries:
            field_values = getattr(entry, "field_values", []) or []
            for d in desc_list:
                child_field_id = int(d["child_field_id"])
                child_sub_key = d["child_sub_field_key"]
                if child_field_id not in child_field_cache:
                    cf_res = await db.execute(
                        select(KPIField)
                        .join(KPI, KPI.id == KPIField.kpi_id)
                        .where(KPIField.id == child_field_id, KPI.organization_id == org_id)
                        .options(selectinload(KPIField.sub_fields))
                    )
                    cf = cf_res.scalar_one_or_none()
                    if cf is not None:
                        child_field_cache[child_field_id] = cf
                child_field = child_field_cache.get(child_field_id)
                if child_field is None:
                    continue
                cache_key = (int(entry.id), int(child_field_id))
                if cache_key not in multi_cache:
                    multi_cache[cache_key] = await _load_multi_line_row_dicts(db, entry_id=entry.id, field=child_field)
                for idx, row in multi_cache[cache_key]:
                    if not isinstance(row, dict):
                        continue
                    cell = row.get(child_sub_key)
                    for label, token in _reverse_ref_tokens_from_cell(cell):
                        if token not in tokens_set:
                            continue
                        token_counts[token] = token_counts.get(token, 0) + 1
                        if include_rows:
                            rows_payload.append(
                                {
                                    "entry_id": entry.id,
                                    "year": entry.year,
                                    "period_key": getattr(entry, "period_key", "") or "",
                                    "value_token": token,
                                    "value_display": label,
                                    "child_field_id": child_field_id,
                                    "child_field_key": d["child_field_key"],
                                    "child_field_name": d["child_field_name"],
                                    "child_sub_field_key": child_sub_key,
                                    "child_sub_field_name": d["child_sub_field_name"],
                                    "row_index": idx,
                                    "row": row,
                                }
                            )

        # Only include child KPI if we found any matching tokens
        if not any((c or 0) > 0 for c in token_counts.values()):
            continue

        values_payload = []
        for token, meta in descriptor_tokens.items():
            count = token_counts.get(token, 0)
            if count <= 0:
                continue
            values_payload.append(
                {
                    "token": token,
                    "label": meta["label"],
                    "count": count,
                }
            )
        # Sort dropdown values by label
        values_payload.sort(key=lambda x: x["label"])

        # Provide a default target field+sub_field for deep-linking into the child multi-line full page.
        default_target = desc_list[0] if desc_list else None
        response_tabs.append(
            {
                "child_kpi_id": child_kpi_id,
                "child_kpi_name": child_kpi_name,
                "child_field_id": default_target.get("child_field_id") if default_target else None,
                "child_sub_field_key": default_target.get("child_sub_field_key") if default_target else None,
                "values": values_payload,
                "rows": rows_payload if include_rows else [],
                "sub_fields": sub_fields_payload if include_rows else [],
            }
        )

    return {
        "time_filter": {
            "year": parent_year,
            "period_key": parent_period_key,
            "effective_time_dimension": effective_td.value if hasattr(effective_td, "value") else str(effective_td),
        },
        "tabs": response_tabs,
    }


@router.get("/for-period", response_model=EntryResponse)
async def get_or_create_entry_for_period(
    kpi_id: int = Query(...),
    year: int = Query(...),
    period_key: str | None = Query(None, description="Period key: '', H1, H2, Q1-Q4, 01-12"),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get entry for the given KPI/year/period; create it (with carry-forward from previous period) if missing. Requires view access."""
    org_id = _org_id(current_user, organization_id)
    can_view = await user_can_view_kpi(db, current_user.id, kpi_id)
    if not can_view:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No access to this KPI")
    pk = (period_key or "").strip()[:8]
    entry, created = await get_or_create_entry(db, current_user.id, org_id, kpi_id, year, period_key=pk)
    if not entry:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="KPI not in organization")
    if created:
        await db.commit()
    await db.refresh(entry, attribute_names=["field_values", "user", "updated_at"])
    return _entry_to_response(entry)


@router.get("/for-period-id", response_model=EntryIdResponse)
async def get_or_create_entry_for_period_id(
    kpi_id: int = Query(...),
    year: int = Query(...),
    period_key: str | None = Query(None, description="Period key: '', H1, H2, Q1-Q4, 01-12"),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Lightweight variant of /entries/for-period that returns only the entry id.
    Avoids loading/serializing field values (used by list pages that only need entry_id).
    """
    org_id = _org_id(current_user, organization_id)
    can_view = await user_can_view_kpi(db, current_user.id, kpi_id)
    if not can_view:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No access to this KPI")
    pk = (period_key or "").strip()[:8]
    entry, created = await get_or_create_entry(db, current_user.id, org_id, kpi_id, year, period_key=pk)
    if not entry:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="KPI not in organization")
    if created:
        await db.commit()
    return EntryIdResponse(id=int(entry.id), created=bool(created))


@router.get("", response_model=list[EntryResponse])
async def list_my_entries(
    kpi_id: int | None = Query(None),
    year: int | None = Query(None),
    period_key: str | None = Query(None, description="Filter by period: '', H1, H2, Q1-Q4, 01-12"),
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List entries for current user; org admin can list all org entries with organization_id."""
    org_id = _org_id(current_user, organization_id)
    as_admin = current_user.role.value in ("ORG_ADMIN", "SUPER_ADMIN")
    entries = await list_entries(
        db, current_user.id, org_id, kpi_id=kpi_id, year=year, period_key=period_key, as_admin=as_admin
    )
    return [_entry_to_response(e) for e in entries]


@router.post("", response_model=EntryResponse, status_code=status.HTTP_201_CREATED)
async def create_or_update_entry(
    body: EntryCreate,
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create or get entry and save values (draft). User must have view access; only values for fields they can edit are saved."""
    org_id = _org_id(current_user, organization_id)
    can_view = await user_can_view_kpi(db, current_user.id, body.kpi_id)
    if not can_view:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not assigned to this KPI")
    # Load KPI with fields and sub_fields so we can allow multi_line_items when user has subfield-level edit
    kpi_res = await db.execute(
        select(KPI)
        .where(KPI.id == body.kpi_id)
        .options(selectinload(KPI.fields).selectinload(KPIField.sub_fields))
    )
    kpi = kpi_res.scalar_one_or_none()
    allowed_values = []
    for v in body.values:
        field = next((x for x in (kpi.fields or []) if x.id == v.field_id), None) if kpi else None
        if field and getattr(field, "field_type", None) == FieldType.multi_line_items:
            can_edit = await user_can_edit_multi_line_field(db, current_user.id, body.kpi_id, field)
        else:
            can_edit = await user_can_edit_field(db, current_user.id, body.kpi_id, v.field_id, None)
        if can_edit:
            allowed_values.append(v)
    if not allowed_values and body.values:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No permission to edit any of the provided fields")
    entry, _ = await get_or_create_entry(
        db, current_user.id, org_id, body.kpi_id, body.year, period_key=body.period_key or ""
    )
    if not entry:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="KPI not in organization")
    try:
        await save_entry_values(db, entry.id, current_user.id, allowed_values, body.kpi_id, org_id)
    except EntryValidationError:
        raise  # Handled by app exception_handler; returns 400 with errors list
    await db.commit()
    await db.refresh(entry, attribute_names=["field_values", "user", "updated_at"])
    return _entry_to_response(entry)


@router.post("/submit", response_model=EntryResponse)
async def submit_entry_route(
    body: EntrySubmit,
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Submit entry (no longer draft). User must have data_entry on KPI or on at least one field."""
    org_id = _org_id(current_user, organization_id)
    ent = await db.execute(select(KPIEntry).where(KPIEntry.id == body.entry_id, KPIEntry.organization_id == org_id))
    entry_row = ent.scalar_one_or_none()
    if not entry_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entry not found")
    can_edit_kpi = await user_can_edit_kpi(db, current_user.id, entry_row.kpi_id)
    field_access = await get_user_field_access_for_kpi(db, current_user.id, entry_row.kpi_id)
    can_edit_any = can_edit_kpi or (field_access is not None and any(p == "data_entry" for p in field_access.values()))
    if not can_edit_any:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to edit this KPI")
    entry = await submit_entry(db, body.entry_id, current_user.id, org_id)
    if not entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entry not found or locked")
    await db.commit()
    await db.refresh(entry, attribute_names=["field_values", "user", "updated_at"])
    return _entry_to_response(entry)


@router.post("/lock", response_model=EntryResponse)
async def lock_entry_route(
  body: EntryLock,
  organization_id: int | None = Query(None),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(require_org_admin),
):
    """Lock or unlock entry (Org Admin)."""
    org_id = _org_id(current_user, organization_id)
    entry = await lock_entry(db, body.entry_id, org_id, body.is_locked)
    if not entry:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Entry not found")
    await db.commit()
    await db.refresh(entry, attribute_names=["field_values", "user", "updated_at"])
    return _entry_to_response(entry)
