"""User API routes (Org Admin)."""

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.core.database import get_db
from app.auth.dependencies import get_current_user, require_org_admin, require_tenant
from app.core.models import User, ExternalUser, UserRole
from app.users.schemas import UserCreate, UserUpdate, UserResponse, UserKpiAssignmentResponse, ExternalUserCreate
from app.users.service import (
    create_user,
    create_external_user,
    get_user,
    get_user_kpi_assignments,
    list_users,
    update_user,
    delete_user,
)

router = APIRouter(prefix="/users", tags=["users"])


def _org_id(user: User, org_id_param: int | None = None) -> int:
    """Resolve organization id for tenant scope. Super Admin may pass org_id_param."""
    if org_id_param is not None and user.role.value == "SUPER_ADMIN":
        return org_id_param
    if user.organization_id is not None:
        return user.organization_id
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Organization required")


@router.get("", response_model=list[UserResponse])
async def list_org_users(
    organization_id: int | None = Query(None, description="Required for Super Admin"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """List users in current organization. Org Admin or Super Admin (with org context)."""
    org_id = _org_id(current_user, organization_id)
    users = await list_users(db, org_id)
    if not users:
        return []
    user_ids = [u.id for u in users]
    ext_res = await db.execute(select(ExternalUser).where(ExternalUser.user_id.in_(user_ids)))
    ext_map = {eu.user_id: eu for eu in ext_res.scalars().all()}
    out: list[UserResponse] = []
    for u in users:
        eu = ext_map.get(u.id)
        out.append(
            UserResponse(
                id=u.id,
                username=u.username,
                email=u.email,
                full_name=u.full_name,
                role=u.role,
                organization_id=u.organization_id,
                is_active=u.is_active,
                description=eu.description if eu else None,
                is_external=eu is not None,
            )
        )
    return out


@router.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def create_org_user(
    body: UserCreate,
    organization_id: int | None = Query(None, description="Required for Super Admin"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """Create user in organization and assign KPIs and report templates."""
    org_id = _org_id(current_user, organization_id)
    user = await create_user(db, org_id, body)
    await db.commit()
    await db.refresh(user)
    return UserResponse.model_validate(user)


@router.post("/external", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
async def create_external_org_user(
    body: ExternalUserCreate,
    organization_id: int | None = Query(None, description="Required for Super Admin"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """
    Create an external user authenticated via external XML-RPC (password is NOT stored/verified internally).
    Org Admin only.
    """
    org_id = _org_id(current_user, organization_id)
    user = await create_external_user(db, org_id, body)
    await db.commit()
    await db.refresh(user)
    return UserResponse.model_validate(user)


@router.get("/external/template")
async def download_external_users_template(
    organization_id: int | None = Query(None, description="Required for Super Admin"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """Download an Excel template for importing external users."""
    from openpyxl import Workbook

    org_id = _org_id(current_user, organization_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "External users"
    ws.append(["username", "full_name", "description", "is_active"])
    # Example row (keeps template user-friendly)
    ws.append(["ext_user_1", "External User 1", "Department / notes", True])

    import uuid

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"external_users_template_{org_id}_{uuid.uuid4().hex[:6]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/external/bulk-upload")
async def upload_external_users_excel(
    append: bool = Query(False, description="If true, append new users only. If false, override existing users in the file."),
    file: UploadFile = File(...),
    organization_id: int | None = Query(None, description="Required for Super Admin"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """Bulk import external users from Excel (.xlsx)."""
    from openpyxl import load_workbook
    from app.core.security import get_password_hash

    org_id = _org_id(current_user, organization_id)
    content = await file.read()

    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty Excel file")

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    header_index: dict[str, int] = {h: i for i, h in enumerate(header) if h}

    def _idx(*names: str) -> int | None:
        for n in names:
            if n in header_index:
                return header_index[n]
        return None

    idx_username = _idx("username")
    idx_full_name = _idx("full_name", "full name", "fullname", "name")
    idx_description = _idx("description", "desc", "notes")
    idx_is_active = _idx("is_active", "active", "isactive")

    if idx_username is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Missing column: username")

    def parse_bool(v: object, default: bool = True) -> bool:
        if v is None or v == "":
            return default
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float)):
            return bool(int(v))
        s = str(v).strip().lower()
        return s in ("1", "true", "yes", "y", "active")

    parsed: list[dict] = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        username = str(r[idx_username] if idx_username is not None else "").strip()
        if not username:
            continue
        def _norm_str(v: object | None) -> str | None:
            if v is None:
                return None
            s = str(v).strip()
            return s if s else None

        full_name = _norm_str(r[idx_full_name]) if idx_full_name is not None else None
        description = _norm_str(r[idx_description]) if idx_description is not None else None
        is_active = parse_bool(r[idx_is_active] if idx_is_active is not None else None, default=True)
        parsed.append({
            "username": username,
            "full_name": full_name,
            "description": description,
            "is_active": is_active,
        })

    if not parsed:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No users found in Excel file")

    # Validate duplicates in file
    seen = set()
    for p in parsed:
        if p["username"] in seen:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Duplicate username in file: {p['username']}")
        seen.add(p["username"])

    usernames = [p["username"] for p in parsed]
    existing_users_res = await db.execute(
        select(User).where(User.organization_id == org_id, User.username.in_(usernames))
    )
    existing_users = existing_users_res.scalars().all()
    existing_by_username = {u.username: u for u in existing_users}

    # External metadata for existing users
    existing_ids = [u.id for u in existing_users]
    ext_res = []
    ext_by_user_id: dict[int, ExternalUser] = {}
    if existing_ids:
        ext_res = await db.execute(select(ExternalUser).where(ExternalUser.user_id.in_(existing_ids)))
        ext_by_user_id = {eu.user_id: eu for eu in ext_res.scalars().all()}

    rows_added = 0
    rows_overridden = 0

    from uuid import uuid4

    for p in parsed:
        username = p["username"]
        if username in existing_by_username:
            if append:
                continue
            user = existing_by_username[username]
            user.full_name = p["full_name"]
            user.is_active = bool(p["is_active"])
            # Ensure external row exists and set description
            eu = ext_by_user_id.get(user.id)
            if eu is None:
                db.add(ExternalUser(user_id=user.id, description=p["description"]))
            else:
                eu.description = p["description"]
            rows_overridden += 1
        else:
            # Create new external user
            dummy_password = f"external:{uuid4().hex}"
            user = User(
                organization_id=org_id,
                username=username,
                email=None,
                full_name=p["full_name"],
                hashed_password=get_password_hash(dummy_password),
                role=UserRole.USER,
                is_active=bool(p["is_active"]),
            )
            db.add(user)
            await db.flush()
            db.add(ExternalUser(user_id=user.id, description=p["description"]))
            rows_added += 1

    await db.commit()
    return {"rows_added": rows_added, "rows_overridden": rows_overridden, "append": append}


@router.get("/{user_id}", response_model=UserResponse)
async def get_org_user(
    user_id: int,
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """Get user by id within organization."""
    org_id = _org_id(current_user, organization_id)
    user = await get_user(db, user_id, org_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    ext_res = await db.execute(select(ExternalUser).where(ExternalUser.user_id == user.id))
    eu = ext_res.scalar_one_or_none()
    return UserResponse(
        id=user.id,
        username=user.username,
        email=user.email,
        full_name=user.full_name,
        role=user.role,
        organization_id=user.organization_id,
        is_active=user.is_active,
        description=eu.description if eu else None,
        is_external=eu is not None,
    )


@router.get("/{user_id}/kpi-assignments", response_model=list[UserKpiAssignmentResponse])
async def get_org_user_kpi_assignments(
    user_id: int,
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """Get user's KPI assignments (kpi_id, permission) within organization."""
    org_id = _org_id(current_user, organization_id)
    assignments = await get_user_kpi_assignments(db, user_id, org_id)
    return [UserKpiAssignmentResponse.model_validate(a) for a in assignments]


@router.patch("/{user_id}", response_model=UserResponse)
async def update_org_user(
    user_id: int,
    body: UserUpdate,
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """Update user and optionally KPI/report assignments."""
    org_id = _org_id(current_user, organization_id)
    user = await update_user(db, user_id, org_id, body)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    await db.commit()
    await db.refresh(user)
    return UserResponse.model_validate(user)


@router.delete("/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_org_user(
    user_id: int,
    organization_id: int | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_org_admin),
):
    """Delete user in organization."""
    org_id = _org_id(current_user, organization_id)
    ok = await delete_user(db, user_id, org_id)
    if not ok:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    await db.commit()
